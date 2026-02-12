import csv
import math
import re
import sys
import uuid
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple, Callable

import cv2
import fitz  # PyMuPDF
import numpy as np
from PySide6.QtCore import QEvent
from PySide6.QtCore import Qt, QPointF, QRectF, QTimer, QElapsedTimer, QObject, QPropertyAnimation, QEasingCurve
from PySide6.QtGui import (
    QImage, QPixmap, QPen, QColor, QBrush, QFont,
    QWheelEvent, QMouseEvent, QCursor
)
from PySide6.QtGui import QKeySequence, QShortcut
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QFileDialog, QWidget, QHBoxLayout, QVBoxLayout,
    QPushButton, QLabel, QGraphicsView, QGraphicsScene, QGraphicsPixmapItem,
    QGraphicsEllipseItem, QGraphicsTextItem, QGraphicsLineItem, QTableWidget, QTableWidgetItem,
    QMessageBox, QInputDialog, QDialog, QFormLayout, QComboBox, QCheckBox,
    QDialogButtonBox, QSpinBox, QListWidget, QListWidgetItem, QAbstractItemView,
    QGroupBox, QSplitter, QSizePolicy, QHeaderView, QTabWidget, QLineEdit, QGridLayout, QSlider, QGraphicsOpacityEffect
)
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


# ---------------------------
# Data Models
# ---------------------------

@dataclass
class ModelSample:
    model: str
    hsv: Tuple[int, int, int]  # representative (H,S,V)

@dataclass
class Dot:
    page_index: int

    # Stable identity for persistence across “modes” + (best-effort) re-detection
    uid: str = field(default_factory=lambda: uuid.uuid4().hex)

    # Stored in PDF coordinate units (points)
    cx_pdf: float = 0.0
    cy_pdf: float = 0.0
    radius_pdf: float = 0.0

    model: str = ""
    excluded: bool = False

    # User-assigned device name (what we’ll show in the NAMES tab)
    name: str = ""

    # Label layout is stored in PDF units too (so exports are stable)
    label_dx_pdf: float = 6.0
    label_dy_pdf: float = -6.0

    # Computed label string (after numbering pass)
    label: Optional[str] = None

    # Useful debug info
    sampled_hsv: Optional[Tuple[int, int, int]] = None

# ---------------------------
# CV Utilities
# ---------------------------

def circularity(cnt) -> float:
    area = cv2.contourArea(cnt)
    if area <= 0:
        return 0.0
    per = cv2.arcLength(cnt, True)
    if per <= 0:
        return 0.0
    return (4.0 * math.pi * area) / (per * per)

def annulus_mean_hsv(hsv_img: np.ndarray, cx: int, cy: int, r: int) -> Optional[Tuple[int, int, int]]:
    """
    Robust ring sampling:
    - Avoids black text in the middle (annulus)
    - Keeps pastels (low S) but rejects dark pixels (low V)
    - Uses "top saturation" pixels to avoid edge/antialias/white bleed
    - Uses circular mean for Hue (critical for hue stability)
    """
    h, w = hsv_img.shape[:2]
    if cx < 0 or cy < 0 or cx >= w or cy >= h:
        return None

    r_outer = max(4, int(r * 0.95))
    r_inner = max(2, int(r * 0.55))

    y1, y2 = max(0, cy - r_outer), min(h, cy + r_outer)
    x1, x2 = max(0, cx - r_outer), min(w, cx + r_outer)

    patch = hsv_img[y1:y2, x1:x2]
    if patch.size == 0:
        return None

    yy, xx = np.ogrid[y1:y2, x1:x2]
    dist2 = (xx - cx) ** 2 + (yy - cy) ** 2
    ring_mask = (dist2 <= r_outer * r_outer) & (dist2 >= r_inner * r_inner)

    ring_pixels = patch[ring_mask]
    if ring_pixels.size == 0:
        return None

    # Reject very dark pixels (black text / outlines)
    ring_pixels = ring_pixels[ring_pixels[:, 2] > 35]
    if ring_pixels.size == 0:
        return None

    # Prefer the "most saturated" pixels in the ring to avoid white bleed / anti-aliasing
    # Keep top 35% saturation pixels (minimum 40 pixels)
    s = ring_pixels[:, 1].astype(np.int32)
    if ring_pixels.shape[0] > 40:
        cutoff = np.percentile(s, 65)  # keep top 35%
        ring_pixels = ring_pixels[s >= cutoff]
        if ring_pixels.size == 0:
            return None

    # Circular mean for hue (0..179)
    hues = ring_pixels[:, 0].astype(np.float32) * (2.0 * math.pi / 180.0)
    mean_sin = np.sin(hues).mean()
    mean_cos = np.cos(hues).mean()
    mean_angle = math.atan2(mean_sin, mean_cos)
    mean_h = int((mean_angle * 180.0 / (2.0 * math.pi)) % 180)

    # Median for S/V is more stable than mean (resists noise)
    med_s = int(np.median(ring_pixels[:, 1]))
    med_v = int(np.median(ring_pixels[:, 2]))

    return mean_h, med_s, med_v

def sample_dot_hsv(hsv_img: np.ndarray, cx: int, cy: int) -> Optional[Tuple[int, int, int]]:
    """
    Adaptive dot sampling for small/large bubbles:
    try multiple ring radii and pick the result with the highest median saturation
    (usually the least contaminated by white bleed / outlines).
    """
    candidates = []
    for r in (6, 8, 10, 12, 14, 16):
        ring = annulus_mean_hsv(hsv_img, cx, cy, r=r)
        if ring is None:
            continue
        h, s, v = ring
        candidates.append((s, v, ring))  # prioritize saturation, then value

    if not candidates:
        return None

    candidates.sort(key=lambda t: (t[0], t[1]), reverse=True)
    return candidates[0][2]


def hue_distance(h1: int, h2: int) -> int:
    # circular hue 0..179
    d = abs(h1 - h2)
    return min(d, 180 - d)

def make_hsv_mask(
    hsv_img: np.ndarray,
    sample_hsv: Tuple[int, int, int],
    dh: int = 10,
    min_s: int = 18,
    min_v: int = 35
) -> np.ndarray:
    """
    Hue window + adaptive S/V floors.

    - Pastels (like lavender C3) have low saturation, so we must not force min_s too high.
    - Vivid colors (like C1) still get protection against washed-out versions (C14).
    """
    H, S, V = sample_hsv

    # Hue band
    h_lo = H - dh
    h_hi = H + dh

    # Adaptive saturation floor:
    # - If the sample itself is vivid, reject washed-out versions aggressively
    # - If the sample itself is pastel, allow low saturation
    if S >= 80:
        s_floor = max(min_s, int(S * 0.70))
    else:
        s_floor = max(min_s, int(S * 0.50))

    # Adaptive value floor to avoid near-white / noise hits
    v_floor = max(min_v, int(V * 0.50))

    s_hi = 255
    v_hi = 255

    if h_lo < 0:
        mask1 = cv2.inRange(hsv_img, (0, s_floor, v_floor), (h_hi, s_hi, v_hi))
        mask2 = cv2.inRange(hsv_img, (179 + h_lo, s_floor, v_floor), (179, s_hi, v_hi))
        return cv2.bitwise_or(mask1, mask2)
    elif h_hi > 179:
        mask1 = cv2.inRange(hsv_img, (h_lo, s_floor, v_floor), (179, s_hi, v_hi))
        mask2 = cv2.inRange(hsv_img, (0, s_floor, v_floor), (h_hi - 179, s_hi, v_hi))
        return cv2.bitwise_or(mask1, mask2)
    else:
        return cv2.inRange(hsv_img, (h_lo, s_floor, v_floor), (h_hi, s_hi, v_hi))

# ADD THIS (new helper)
def peaks_from_distance_transform(mask_u8: np.ndarray) -> List[Tuple[int, int, int]]:
    """
    Given a binary mask (0/255), find dot centers via distance-transform peaks.
    Returns list of (cx, cy, r_px_est).
    Works much better than contours when two dots touch/merge into one component.
    """
    if mask_u8 is None or mask_u8.size == 0:
        return []

    # Ensure binary 0/255
    m = (mask_u8 > 0).astype(np.uint8) * 255

    # Distance to nearest zero pixel (background)
    dist = cv2.distanceTransform(m, cv2.DIST_L2, 5)

    if dist.max() <= 0:
        return []

    # Local maxima of distance map
    dist_dil = cv2.dilate(dist, np.ones((3, 3), np.uint8))
    peaks = (dist == dist_dil)

    # Minimum radius in pixels to consider a "dot"
    # (tune if you need: 2.0–4.0 are typical)
    peaks &= (dist >= 2.5)

    ys, xs = np.where(peaks)
    if len(xs) == 0:
        return []

    # Sort peaks by strength (radius) desc
    candidates = sorted(
        [(int(x), int(y), float(dist[int(y), int(x)])) for x, y in zip(xs, ys)],
        key=lambda t: t[2],
        reverse=True
    )

    kept: List[Tuple[int, int, int]] = []
    for cx, cy, r_f in candidates:
        r = max(2, int(round(r_f)))

        # NMS: don't keep two peaks that are basically the same dot
        # min distance scales with r so close dots can still both survive
        min_sep = max(3, int(round(r * 1.4)))

        too_close = False
        for kx, ky, kr in kept:
            if math.hypot(cx - kx, cy - ky) < min_sep:
                too_close = True
                break
        if too_close:
            continue

        kept.append((cx, cy, r))

    return kept

def filter_peaks_by_component_stats(
    mask_u8: np.ndarray,
    peaks: List[Tuple[int, int, int]],
    comp_area_min: int = 12,
    comp_area_max: int = 2600,
    extent_min: float = 0.35,
    aspect_max: float = 3.2
) -> List[Tuple[int, int, int]]:
    """
    Filters DT peaks by the connected-component they fall inside.
    IMPORTANT: does NOT dedupe by component label (close/touching dots can share a blob
    but still produce multiple DT peaks — we keep them).

    Rejects:
    - tiny specks (area too small)
    - long skinny strokes (aspect too large)
    - very unfilled bbox (extent too low)
    """
    if mask_u8 is None or mask_u8.size == 0 or not peaks:
        return peaks

    m = (mask_u8 > 0).astype(np.uint8) * 255
    _, labels, stats, _ = cv2.connectedComponentsWithStats(m, connectivity=8)

    h, w = m.shape[:2]
    out: List[Tuple[int, int, int]] = []

    for cx, cy, r in peaks:
        if cx < 0 or cy < 0 or cx >= w or cy >= h:
            continue

        lab = int(labels[cy, cx])
        if lab <= 0:
            continue

        area = int(stats[lab, cv2.CC_STAT_AREA])
        bw = int(stats[lab, cv2.CC_STAT_WIDTH])
        bh = int(stats[lab, cv2.CC_STAT_HEIGHT])

        if area < comp_area_min or area > comp_area_max:
            continue

        bbox_area = max(1, bw * bh)
        extent = area / bbox_area
        if extent < extent_min:
            continue

        aspect = max(bw, bh) / max(1, min(bw, bh))
        if aspect > aspect_max:
            continue

        out.append((cx, cy, r))

    return out


def detect_dots_for_model(
    bgr: np.ndarray,
    model_sample: ModelSample,
    area_min: int = 20,
    area_max: int = 8000,
    circ_min: float = 0.65,
    dh: int = 12,
    min_s: int = 70,
    min_v: int = 70
) -> List[Tuple[int, int, int, Tuple[int, int, int]]]:
    """
    Returns list of (cx, cy, radius, ringHSV) in pixel coordinates for dots matching this sample.
    """
    hsv = cv2.cvtColor(bgr, cv2.COLOR_BGR2HSV)
    mask = make_hsv_mask(
        hsv,
        model_sample.hsv,
        dh=dh,
        min_s=min_s,
        min_v=min_v
    )

    # cleanup (preserve tiny/pastel dots)
    # cleanup:
    # Pastel dots are low-S and often close together — CLOSE can "bridge" gaps and merge dots.
    mask = cv2.medianBlur(mask, 3)

    sample_s = int(model_sample.hsv[1])
    if sample_s >= 60:
        # Vivid colors: CLOSE is helpful (fills tiny holes)
        mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, np.ones((3, 3), np.uint8), iterations=1)
    else:
        # Pastels: avoid CLOSE so neighboring dots don't fuse
        # (Optional tiny OPEN if you get speckle noise; keep it very light)
        # mask = cv2.morphologyEx(mask, cv2.MORPH_OPEN, np.ones((2, 2), np.uint8), iterations=1)
        mask = cv2.morphologyEx(mask, cv2.MORPH_OPEN, np.ones((2, 2), np.uint8), iterations=1)

    results: List[Tuple[int, int, int, Tuple[int, int, int]]] = []

    # 1) Preferred: distance-transform peaks (splits merged blobs)
    peaks = peaks_from_distance_transform(mask)
    peaks = filter_peaks_by_component_stats(mask, peaks, comp_area_min=12, comp_area_max=2600, extent_min=0.35, aspect_max=3.2)

    for cx, cy, rr in peaks:
        area_est = math.pi * (rr * rr)
        if area_est < area_min or area_est > area_max:
            continue

        ring = annulus_mean_hsv(hsv, cx, cy, rr)
        if ring is None:
            continue

        results.append((cx, cy, rr, ring))

    # 2) Fallback: contours (if peaks found nothing)
    if not results:
        contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        for cnt in contours:
            area = cv2.contourArea(cnt)
            if area < area_min or area > area_max:
                continue

            circ = circularity(cnt)
            if circ < circ_min:
                continue

            (x, y), r = cv2.minEnclosingCircle(cnt)
            cx, cy, rr = int(x), int(y), int(r)

            ring = annulus_mean_hsv(hsv, cx, cy, rr)
            if ring is None:
                continue

            results.append((cx, cy, rr, ring))

    return results



# ---------------------------
# Graphics Items
# ---------------------------

class DotEllipseItem(QGraphicsEllipseItem):
    """
    A clickable dot overlay ellipse.
    """
    def __init__(self, dot_id: int, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.dot_id = dot_id
        self.setFlag(QGraphicsEllipseItem.ItemIsSelectable, True)

class LabelTextItem(QGraphicsTextItem):
    """
    Draggable label with dot association + background + colored border.
    Supports dynamic scaling (font + padding + border thickness).
    Calls back to the app whenever it moves so connectors can update live.
    """
    def __init__(
        self,
        dot_id: int,
        text: str,
        base_font_pt: int = 14,
        scale: float = 1.0,
        border_color: Optional[QColor] = None,
        on_moved: Optional[Callable[[int], None]] = None
    ):
        super().__init__(text)
        self.dot_id = dot_id
        self._on_moved = on_moved

        self.setFlag(QGraphicsTextItem.ItemIsMovable, True)
        self.setFlag(QGraphicsTextItem.ItemIsSelectable, True)

        self.text_color = QColor(0, 0, 0)
        self.bg_color = QColor(255, 255, 255)
        self.border_color = border_color if border_color is not None else QColor(0, 0, 0)

        # Base metrics (unscaled)
        self._base_font_pt = int(base_font_pt)
        self._base_padding = 3
        self._base_border_w = 2

        # Scaled metrics (computed)
        self.padding = self._base_padding
        self.border_w = self._base_border_w

        self.setDefaultTextColor(self.text_color)
        self.set_scale(scale)

    def set_border_color(self, c: QColor):
        self.border_color = QColor(c)
        self.update()

    def set_scale(self, scale: float):
        # clamp to avoid 0 or negative font sizes
        s = max(0.05, float(scale))
        font_pt = max(1, int(round(self._base_font_pt * s)))

        self.padding = max(0, int(round(self._base_padding * s)))
        self.border_w = max(1, int(round(self._base_border_w * s)))

        self.setFont(QFont("Arial", font_pt))

        # Force geometry recalculation so the scene repaints correctly
        self.prepareGeometryChange()
        self.update()

    def boundingRect(self):
        # Expand the bounding rect to include our painted padding/background/border.
        # Prevents trails/artifacting during drag.
        r = super().boundingRect()
        return r.adjusted(-self.padding, -self.padding, self.padding, self.padding)

    def itemChange(self, change, value):
        # Update connector line live while moving
        if change == QGraphicsTextItem.ItemPositionHasChanged:
            if callable(self._on_moved):
                try:
                    self._on_moved(self.dot_id)
                except Exception:
                    pass
        return super().itemChange(change, value)

    def paint(self, painter, option, widget=None):
        rect = super().boundingRect().adjusted(-self.padding, -self.padding, self.padding, self.padding)
        painter.setBrush(QBrush(self.bg_color))
        painter.setPen(QPen(self.border_color, self.border_w))
        painter.drawRoundedRect(rect, 2, 2)

        super().paint(painter, option, widget)

class DotEditDialog(QDialog):
    def __init__(self, parent, dot: Dot, model_list: List[str]):
        super().__init__(parent)
        self.setWindowTitle("Edit Dot")
        self.dot = dot
        self.model_list = model_list

        layout = QFormLayout(self)

        # --- Dropdown (existing sampled/known models) ---
        self.model_combo = QComboBox()
        self.model_combo.addItems(model_list)
        if dot.model in model_list:
            self.model_combo.setCurrentText(dot.model)
        layout.addRow("Model", self.model_combo)

        # --- NEW: per-dot custom model override ---
        # This allows typing a model that hasn't been sampled yet.
        # It applies ONLY to this dot (no global forcing).
        self.model_custom = QLineEdit()
        self.model_custom.setPlaceholderText("Type a model not in the list (optional)")
        # Pre-fill if current dot.model is NOT in the list, or just show it for convenience.
        if dot.model and dot.model not in model_list:
            self.model_custom.setText(dot.model)
        layout.addRow("Model (custom)", self.model_custom)

        self.excluded_cb = QCheckBox("Exclude (ignore this dot)")
        self.excluded_cb.setChecked(dot.excluded)
        layout.addRow("", self.excluded_cb)

        if dot.sampled_hsv:
            layout.addRow("Sampled HSV", QLabel(str(dot.sampled_hsv)))

        self.delete_cb = QCheckBox("Delete dot entirely")
        layout.addRow("", self.delete_cb)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addRow(buttons)

    def result_values(self):
        # Prefer the custom model field if the user typed anything.
        custom = (self.model_custom.text() or "").strip()
        chosen = (self.model_combo.currentText() or "").strip()
        model = custom if custom else chosen

        return {
            "model": model,
            "excluded": self.excluded_cb.isChecked(),
            "delete": self.delete_cb.isChecked()
        }


class ExportImagesDialog(QDialog):
    def __init__(self, parent):
        super().__init__(parent)
        self.setWindowTitle("Export Images")

        layout = QFormLayout(self)
        self.dpi_spin = QSpinBox()
        self.dpi_spin.setRange(72, 2400)
        self.dpi_spin.setValue(600)
        self.dpi_spin.setSingleStep(50)
        layout.addRow("DPI", self.dpi_spin)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addRow(buttons)

    def dpi(self) -> int:
        return int(self.dpi_spin.value())
def to_proper_case(s: str) -> str:
    """
    Convert ALL CAPS / messy text into human-friendly Proper Case.
    - Preserves apostrophes nicely: MEN'S -> Men's, KIDS' -> Kids'
    - Capitalizes after spaces, hyphens, and slashes.
    """
    s = " ".join((s or "").strip().split())
    s = s.lower()
    return re.sub(r'(^|[\s\-/])([a-z])', lambda m: m.group(1) + m.group(2).upper(), s)
class DeviceNamingDialog(QDialog):
    """
    Wizard-style naming dialog:
    - Prev/current/next summary in corner
    - Name preview
    - Custom name field
    - Tags/modifiers including '#'
    - Button to capture text by drag-selecting on the map
    - Next saves and advances
    """
    def __init__(self, parent, dot_uids: List[str]):
        super().__init__(parent)
        self.setWindowTitle("Device Naming")
        self.parent_app: PdfDotLabeler = parent  # type: ignore
        self.dot_uids = dot_uids
        self.idx = 0
        # Docking behavior
        self._did_initial_dock = False
        self._user_moved = False

        # UI state
        self.selected_text = ""
        self.tags = set()
        self.setWindowFlags(self.windowFlags() | Qt.WindowStaysOnTopHint)

        layout = QVBoxLayout(self)

        # (Removed Prev/Current/Next display block)

        # Selected text + custom name
        form = QFormLayout()
        self.lbl_selected = QLabel("")
        self.edit_custom = QLineEdit()
        self.lbl_preview = QLabel("")
        self.lbl_preview.setStyleSheet("font-weight: bold;")

        form.addRow("Selected text:", self.lbl_selected)
        form.addRow("Custom name:", self.edit_custom)
        form.addRow("Preview:", self.lbl_preview)
        layout.addLayout(form)

        # Tag buttons
        tag_row = QHBoxLayout()
        self.btn_tag_hash = QPushButton("# (Ctrl+3)")
        self.btn_tag_in = QPushButton("In")
        self.btn_tag_out = QPushButton("Out")
        self.btn_tag_entrance = QPushButton("Entrance")
        self.btn_tag_exit = QPushButton("Exit")
        self.btn_tag_room = QPushButton("Room")
        self.btn_tag_closet = QPushButton("Closet")

        for b in [self.btn_tag_hash, self.btn_tag_in, self.btn_tag_out, self.btn_tag_entrance,
                  self.btn_tag_exit, self.btn_tag_room, self.btn_tag_closet]:
            b.setCheckable(True)
            tag_row.addWidget(b)
        # Keep a reusable list of tag buttons (used for persistence)
        self.tag_buttons = [
            self.btn_tag_hash,
            self.btn_tag_in,
            self.btn_tag_out,
            self.btn_tag_entrance,
            self.btn_tag_exit,
            self.btn_tag_room,
            self.btn_tag_closet
        ]

        # Persist last-entered custom text + tag states across Next/Prev
        self.last_custom_text: str = ""
        self.last_tag_states = {str(i): False for i, _ in enumerate(self.tag_buttons)}

        # Zoom hotkeys that work even while this dialog has focus
        self.sc_zoom_in = QShortcut(QKeySequence("Ctrl+="), self)
        self.sc_zoom_in.activated.connect(lambda: self.parent_app.view.scale(1.15, 1.15))
        # Shortcut: Ctrl+3 toggles the "# (series)" tag
        self.short_tag_hash = QShortcut(QKeySequence("Ctrl+3"), self)
        self.short_tag_hash.setContext(Qt.ApplicationShortcut)
        self.short_tag_hash.activated.connect(lambda: self.btn_tag_hash.toggle())

        self.sc_zoom_in2 = QShortcut(QKeySequence("Ctrl++"), self)
        self.sc_zoom_in2.activated.connect(lambda: self.parent_app.view.scale(1.15, 1.15))

        self.sc_zoom_out = QShortcut(QKeySequence("Ctrl+-"), self)
        self.sc_zoom_out.activated.connect(lambda: self.parent_app.view.scale(1 / 1.15, 1 / 1.15))

        self.sc_zoom_reset = QShortcut(QKeySequence("Ctrl+0"), self)
        self.sc_zoom_reset.activated.connect(
            lambda: self.parent_app.focus_dot(self.current_dot().uid) if self.current_dot() else None)

        layout.addLayout(tag_row)

        # Buttons row 1: Capture + Clear (left)
        btn_row1 = QHBoxLayout()
        self.btn_capture = QPushButton("Capture Text (Ctrl+S)")
        self.btn_clear = QPushButton("Clear (Ctrl+D)")
        btn_row1.addWidget(self.btn_capture)
        btn_row1.addWidget(self.btn_clear)
        btn_row1.addStretch(1)
        layout.addLayout(btn_row1)

        # Buttons row 2: Previous/Next centered, Close bottom-right
        btn_row2 = QHBoxLayout()
        self.btn_prev = QPushButton("Previous (Ctrl+E)")
        self.btn_next = QPushButton("Next (Ctrl+R)")
        self.btn_close = QPushButton("Close")

        # Double height for nav/close buttons
        for b in (self.btn_prev, self.btn_next, self.btn_close):
            b.setMinimumHeight(56)

        btn_row2.addStretch(1)
        btn_row2.addWidget(self.btn_prev)
        btn_row2.addWidget(self.btn_next)
        btn_row2.addStretch(1)
        btn_row2.addWidget(self.btn_close)
        layout.addLayout(btn_row2)

        # Wiring
        self.btn_capture.clicked.connect(self.capture_text)
        self.btn_clear.clicked.connect(self.clear_all)
        self.btn_prev.clicked.connect(self.go_prev)
        self.btn_next.clicked.connect(self.save_and_next)
        self.btn_close.clicked.connect(self.accept)
        # Hotkeys
        # --- Keyboard shortcuts (work even when typing in the name field) ---
        self.short_capture = QShortcut(QKeySequence("Ctrl+S"), self)
        self.short_capture.setContext(Qt.ApplicationShortcut)
        self.short_capture.activated.connect(self.capture_text)

        self.short_clear = QShortcut(QKeySequence("Ctrl+D"), self)
        self.short_clear.setContext(Qt.ApplicationShortcut)
        self.short_clear.activated.connect(self.clear_all)

        self.short_prev = QShortcut(QKeySequence("Ctrl+E"), self)
        self.short_prev.setContext(Qt.ApplicationShortcut)
        self.short_prev.activated.connect(self.go_prev)

        self.short_next = QShortcut(QKeySequence("Ctrl+R"), self)
        self.short_next.setContext(Qt.ApplicationShortcut)
        self.short_next.activated.connect(self.save_and_next)

        # QLineEdit can consume Ctrl+D etc.; event-filter it to guarantee shortcuts work
        self.edit_custom.installEventFilter(self)

        self.edit_custom.textChanged.connect(self.update_preview)

        for b in [self.btn_tag_hash, self.btn_tag_in, self.btn_tag_out, self.btn_tag_entrance,
                  self.btn_tag_exit, self.btn_tag_room, self.btn_tag_closet]:
            b.toggled.connect(self.update_preview)

        # Resume from last uid if available
        if self.parent_app.naming_last_uid and self.parent_app.naming_last_uid in self.dot_uids:
            self.idx = self.dot_uids.index(self.parent_app.naming_last_uid)

        self.refresh_context()

    def _dock_bottom_left(self):
        if self._user_moved:
            return

        self.adjustSize()

        # Dock relative to the MAIN WINDOW (bottom-left), not the entire screen
        main = self.parent_app
        main_geo = main.frameGeometry()

        screen = main.screen() or QApplication.primaryScreen()
        work = screen.availableGeometry()

        margin = 14
        x = main_geo.left() + margin
        y = main_geo.bottom() - self.height() - margin

        # Clamp to the visible work area so it never goes off-screen
        x = max(work.left() + margin, min(x, work.right() - self.width() - margin))
        y = max(work.top() + margin, min(y, work.bottom() - self.height() - margin))

        self.move(x, y)

    def current_dot(self) -> Optional[Dot]:
        if not self.dot_uids:
            return None
        uid = self.dot_uids[self.idx]
        return self.parent_app.get_dot_by_uid(uid)

    def showEvent(self, event):
        super().showEvent(event)

        # Only dock once, and only if the user hasn't moved the window
        if self._did_initial_dock or self._user_moved:
            return

        self._did_initial_dock = True

        # Let the window manager finish placing the dialog, then dock it.
        QTimer.singleShot(0, self._dock_bottom_left)

    def moveEvent(self, event):
        # Only treat moves as "user moved it" AFTER we have performed the initial dock.
        if self.isVisible() and self._did_initial_dock:
            self._user_moved = True
        super().moveEvent(event)

    def refresh_context(self):
        d = self.current_dot()
        if not d:
            return

        # move map + halo (already zooms into the dot area via focus_dot())
        self.parent_app.focus_dot(d.uid)
        # Keep NAMES tab focused + highlight/center-scroll current device
        self.parent_app.highlight_names_uid(d.uid)

        # reset per-dot selected text, but KEEP the last typed custom text
        self.selected_text = ""
        self.lbl_selected.setText("")

        # Restore previous tag states (so # stays enabled across Next/Prev)
        for i, b in enumerate(self.tag_buttons):
            b.blockSignals(True)
            b.setChecked(bool(self.last_tag_states.get(str(i), False)))
            b.blockSignals(False)

        # Restore last custom text into the field
        self.edit_custom.blockSignals(True)
        self.edit_custom.setText(self.last_custom_text)
        self.edit_custom.blockSignals(False)

        self.update_preview()

        # Highlight by default so typing overwrites immediately
        self.edit_custom.setFocus(Qt.TabFocusReason)
        self.edit_custom.selectAll()

    def capture_text(self):
        # Hide the dialog so it can't steal focus while the user drags on the map
        self.hide()

        def _got_text(txt: str):
            pretty = to_proper_case(txt or "")

            # Store + display Selected text in Proper Case
            self.selected_text = pretty
            self.lbl_selected.setText(pretty)

            # ALSO auto-populate Custom name (Proper Case)
            self.edit_custom.blockSignals(True)
            self.edit_custom.setText(pretty)
            self.edit_custom.blockSignals(False)
            self.edit_custom.selectAll()

            self.update_preview()

            # Bring the dialog back after selection completes
            self.show()
            self.raise_()
            self.activateWindow()

        self.parent_app.raise_()
        self.parent_app.activateWindow()
        self.parent_app.enter_text_select_mode(_got_text)

    def clear_all(self):
        self.selected_text = ""
        self.lbl_selected.setText("")
        self.edit_custom.setText("")
        for b in [self.btn_tag_hash, self.btn_tag_in, self.btn_tag_out, self.btn_tag_entrance,
                  self.btn_tag_exit, self.btn_tag_room, self.btn_tag_closet]:
            b.setChecked(False)
        self.update_preview()

    def compute_preview_name(self) -> str:
        base = self.edit_custom.text().strip()
        if not base:
            base = self.selected_text.strip()

        if not base:
            return ""

        suffixes = []
        if self.btn_tag_in.isChecked(): suffixes.append("In")
        if self.btn_tag_out.isChecked(): suffixes.append("Out")
        if self.btn_tag_entrance.isChecked(): suffixes.append("Entrance")
        if self.btn_tag_exit.isChecked(): suffixes.append("Exit")
        if self.btn_tag_room.isChecked(): suffixes.append("Room")
        if self.btn_tag_closet.isChecked(): suffixes.append("Closet")

        # Apply suffix tags
        if suffixes:
            base = base + " " + " ".join(suffixes)

        # Series tag (#): assign incrementing suffix number for identical base names
        # Series tag (#): assign incrementing suffix number for identical base names
        # Matches: "Base", "Base 1", "Base 2", ... (case-insensitive)
        if self.btn_tag_hash.isChecked():
            key = base.strip()
            key_l = key.lower()

            # count across *all* visible dots (not just ones with labels)
            pat = re.compile(rf"^{re.escape(key_l)}(?:\s+(\d+))?$", re.IGNORECASE)

            max_n = 0
            saw_base_without_number = False

            for pi in self.parent_app.visible_pages:
                for dd in self.parent_app.dots_by_page.get(pi, []):
                    if dd.excluded:
                        continue
                    nm = (dd.name or "").strip()
                    if not nm:
                        continue

                    m = pat.match(nm.lower())
                    if not m:
                        continue

                    if m.group(1):
                        try:
                            max_n = max(max_n, int(m.group(1)))
                        except Exception:
                            pass
                    else:
                        # exact "Base" with no trailing number
                        saw_base_without_number = True

            # next number:
            # - if we have Base 1..N => N+1
            # - else if we only have Base (no number) => 2
            # - else => 1
            if max_n > 0:
                next_n = max_n + 1
            elif saw_base_without_number:
                next_n = 2
            else:
                next_n = 1

            base = f"{key} {next_n}"

        return base.strip()

    def update_preview(self):
        preview = self.compute_preview_name()
        self.lbl_preview.setText(preview if preview else "_____")

    def save_current(self):
        d = self.current_dot()
        if not d:
            return

        # Persist what the user typed so it carries forward to the next dot
        self.last_custom_text = self.edit_custom.text()

        # Persist tag button states (so # stays enabled across Next/Prev)
        self.last_tag_states = {str(i): b.isChecked() for i, b in enumerate(self.tag_buttons)}

        new_name = self.compute_preview_name().strip()
        if new_name:
            d.name = new_name

        # persist resume cursor
        self.parent_app.naming_last_uid = d.uid

        # refresh tables + names tab
        self.parent_app.refresh_totals_ui()
        self.parent_app.highlight_names_uid(d.uid)

    def save_and_next(self):
        self.save_current()
        if self.idx < len(self.dot_uids) - 1:
            self.idx += 1
            self.refresh_context()
        else:
            # end
            self.parent_app.statusBar().showMessage("Naming complete (last device).")

    def go_prev(self):
        if self.idx > 0:
            self.save_current()
            self.idx -= 1
            self.refresh_context()
    def eventFilter(self, obj, event):
        if obj is self.edit_custom and event.type() == QEvent.Type.KeyPress:
            # Make these work even when the cursor is in the name field
            if event.modifiers() == Qt.ControlModifier and event.key() == Qt.Key_S:
                self.capture_text()
                return True
            if event.modifiers() == Qt.ControlModifier and event.key() == Qt.Key_D:
                self.clear_all()
                return True
            if event.modifiers() == Qt.ControlModifier and event.key() == Qt.Key_E:
                self.go_prev()
                return True
            if event.modifiers() == Qt.ControlModifier and event.key() == Qt.Key_R:
                self.save_and_next()
                return True

        return super().eventFilter(obj, event)


# ---------------------------
# Main App
# ---------------------------
class TotalsTabWidget(QWidget):
    """
    Container widget for a totals tab:
      - main_table: scrollable per-model rows
      - footer_table: 1-row frozen TOTAL bar (does not scroll)
    """
    def __init__(self, main_table: QTableWidget, footer_table: QTableWidget):
        super().__init__()
        self.main_table = main_table
        self.footer_table = footer_table

        lay = QVBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)

        lay.addWidget(self.main_table, 1)
        lay.addWidget(self.footer_table, 0)

class ModelReorderTable(QTableWidget):
    """
    QTableWidget that ONLY allows row reordering ABOVE/BELOW other rows.
    It never allows "drop on" an item/cell.
    """
    def __init__(self, *args, on_reorder=None, **kwargs):
        super().__init__(*args, **kwargs)
        self._on_reorder = on_reorder

        self.setDragEnabled(True)
        self.setAcceptDrops(True)
        self.setDropIndicatorShown(True)
        self.setDragDropOverwriteMode(False)
        self.setDragDropMode(QAbstractItemView.DragDrop)

        self.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.setSelectionMode(QAbstractItemView.SingleSelection)

    def _move_row(self, src_row: int, dest_row: int):
        """
        Move src_row to dest_row (dest_row is the INSERT position).
        """
        if src_row < 0:
            return

        cols = self.columnCount()

        # Extract items
        items = [self.takeItem(src_row, c) for c in range(cols)]
        row_h = self.rowHeight(src_row)

        # Remove the row
        self.removeRow(src_row)

        # Adjust destination because we removed a row above it
        if dest_row > src_row:
            dest_row -= 1

        dest_row = max(0, min(dest_row, self.rowCount()))

        # Insert and restore items
        self.insertRow(dest_row)
        self.setRowHeight(dest_row, row_h)

        for c, it in enumerate(items):
            if it is not None:
                self.setItem(dest_row, c, it)

        # Keep selection stable
        self.setCurrentCell(dest_row, 0)
        self.selectRow(dest_row)

    def dragMoveEvent(self, event):
        # Force Qt to show only Above/Below indicators (not "OnItem")
        pos = event.position().toPoint()
        row = self.rowAt(pos.y())

        if row >= 0:
            rect = self.visualRect(self.model().index(row, 0))
            # If cursor is in the middle band, Qt might treat it as OnItem.
            # We still accept the drag, but we'll convert it to above/below at drop time.
            event.acceptProposedAction()
        else:
            event.acceptProposedAction()

    def dropEvent(self, event):
        if event.source() != self:
            super().dropEvent(event)
            return

        src = self.currentRow()
        if src < 0:
            event.ignore()
            return

        pos = event.position().toPoint()
        row = self.rowAt(pos.y())

        # If dropped below all rows, treat as append
        if row < 0:
            dest = self.rowCount()
        else:
            rect = self.visualRect(self.model().index(row, 0))
            # Decide ABOVE vs BELOW based on cursor Y relative to row midpoint
            if pos.y() < rect.center().y():
                dest = row
            else:
                dest = row + 1

        # If no real move, ignore
        if dest == src or dest == src + 1:
            event.ignore()
            return

        self.setUpdatesEnabled(False)
        try:
            self._move_row(src, dest)
        finally:
            self.setUpdatesEnabled(True)

        # Notify parent to sync model_order
        if callable(self._on_reorder):
            self._on_reorder()

        event.acceptProposedAction()

class PdfOpenDialog(QFileDialog):
    def __init__(self, parent=None):
        super().__init__(parent)

        self.setFileMode(QFileDialog.ExistingFile)
        self.setNameFilter("PDF Files (*.pdf)")
        self.setViewMode(QFileDialog.Detail)

        # Ensures keys go to the dialog immediately
        self.setFocusPolicy(Qt.StrongFocus)
        # Spacebar: if a file is selected, accept/open
        self._sc_space = QShortcut(QKeySequence(Qt.Key_Space), self)
        self._sc_space.setContext(Qt.WidgetWithChildrenShortcut)
        self._sc_space.activated.connect(self._accept_if_selected)

        # Helpful default directory: current PDF folder if available
        try:
            if parent and getattr(parent, "pdf_path", None):
                import os
                self.setDirectory(os.path.dirname(parent.pdf_path))
        except Exception:
            pass
    def _accept_if_selected(self):
        # Only accept if there is an actual selected file
        files = self.selectedFiles()
        if files:
            self.accept()


    def showEvent(self, event):
        super().showEvent(event)

        # Force the dialog + its file view to actually take keyboard focus.
        self.activateWindow()
        self.raise_()
        self.setFocus(Qt.ActiveWindowFocusReason)

        # Try to focus the file list / tree view so arrows + space work immediately
        for w in self.findChildren(QWidget):
            name = w.metaObject().className()
            if name in ("QListView", "QTreeView"):
                w.setFocus(Qt.ActiveWindowFocusReason)
                break

    def keyPressEvent(self, event):
        # Spacebar should act like "Open" when a file is selected
        if event.key() == Qt.Key_Space:
            if self.selectedFiles():
                self.accept()
                return
        super().keyPressEvent(event)

# ADD (NEW): Dot size preview widget (marching-ants circles under Dot Size section)
class DotSizePreviewWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._min_r = 4
        self._exp_r = 8
        self._max_r = 12
        self._dash_offset = 0.0

        self.setMinimumHeight(70)
        self.setMaximumHeight(90)

    def set_radii(self, min_r: float, exp_r: float, max_r: float):
        self._min_r = float(min_r)
        self._exp_r = float(exp_r)
        self._max_r = float(max_r)
        self.update()

    def set_dash_offset(self, offset: float):
        self._dash_offset = float(offset)
        self.update()

    def paintEvent(self, event):
        from PySide6.QtGui import QPainter, QPen
        from PySide6.QtCore import Qt, QPointF

        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing, True)

        w = self.width()
        h = self.height()
        cx = w * 0.5
        cy = h * 0.5
        center = QPointF(cx, cy)

        # Keep it visually compact regardless of the actual radius values:
        # Scale all three radii so the max circle fits the widget nicely.
        pad = 8.0
        max_draw_r = min(cx, cy) - pad
        src_max = max(self._max_r, 1.0)
        scale = max_draw_r / src_max

        def draw_circle(radius, alpha):
            r = radius * scale
            pen = QPen(QColor(255, 255, 255, int(255 * alpha)))
            pen.setWidth(2)
            pen.setStyle(Qt.CustomDashLine)
            pen.setDashPattern([4, 4])  # marching ants
            pen.setDashOffset(self._dash_offset)
            p.setPen(pen)
            p.setBrush(Qt.NoBrush)
            p.drawEllipse(center, r, r)

        # Min/Max at 50% opacity, Expected at 100% (emphasized)
        draw_circle(self._min_r, 0.5)
        draw_circle(self._max_r, 0.5)
        draw_circle(self._exp_r, 1.0)

        p.end()

class PdfDotLabeler(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Camera Plan Dot Labeler (GUI)")

        # PDF state
        self.doc: Optional[fitz.Document] = None
        self.pdf_path: Optional[str] = None

        # Render/display zoom (multiplier to PDF points)
        self.render_zoom: float = 2.5

        # Page layout in scene
        self.page_y_offsets: List[float] = []
        self.page_sizes_px: List[Tuple[int, int]] = []
        self.page_pixmaps: List[QGraphicsPixmapItem] = []

        # CV samples and model ordering
        self.samples: Dict[str, ModelSample] = {}
        self.model_order: List[str] = []  # drag/drop ordering

        # Page management (remove pages from program without touching the PDF file)
        self.removed_pages: set[int] = set()  # doc page indexes that are "deleted" in the app
        self.visible_pages: List[int] = []  # ordered list of doc page indexes currently shown
        self.doc_to_visible_index: Dict[int, int] = {}  # doc page -> visible index mapping

        # Renameable floor titles (default = "Floor {n}")
        self.page_titles: Dict[int, str] = {}  # doc page index -> title text

        # Dots stored in PDF coords
        self.dots_by_page: Dict[int, List[Dot]] = {}

        # Exclusion zones stored in PDF coords per page
        self.exclusion_zones_by_page: Dict[int, List[QRectF]] = {}

        # Exclusion drawing mode state
        self.mode_exclusion_zone = False
        self.exclusion_drag_start_scene: Optional[QPointF] = None
        self.exclusion_drag_rect_item = None

        # Graphics scene / overlays
        self.scene = QGraphicsScene()
        self.view = QGraphicsView(self.scene)

        # FIX: prevent zoom/pan "trail" artifacts with scaled/AA overlays
        self.view.setViewportUpdateMode(QGraphicsView.FullViewportUpdate)
        self.view.setCacheMode(QGraphicsView.CacheNone)

        # Better pan/zoom feel
        self.view.setTransformationAnchor(QGraphicsView.AnchorUnderMouse)
        self.view.setResizeAnchor(QGraphicsView.AnchorUnderMouse)
        self.view.setDragMode(QGraphicsView.ScrollHandDrag)

        self.view.setFocusPolicy(Qt.StrongFocus)
        self.view.viewport().setFocusPolicy(Qt.StrongFocus)
        self.view.viewport().setFocus()

        self.view.viewport().installEventFilter(self)
        # Transient banner overlay shown at the top of the viewport (e.g., on load)
        self._banner = QLabel(self.view.viewport())
        self._banner.hide()
        self._banner.setWordWrap(True)
        self._banner.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        self._banner.setStyleSheet(
            "QLabel {"
            "  background-color: rgba(200, 255, 200, 230);"
            "  color: black;"
            "  font-weight: bold;"
            "  border: 1px solid rgba(0,0,0,50);"
            "  border-radius: 8px;"
            "  padding: 8px 10px;"
            "}"
        )
        self._banner.setAttribute(Qt.WA_TransparentForMouseEvents, True)
        self._banner_effect = QGraphicsOpacityEffect(self._banner)
        self._banner.setGraphicsEffect(self._banner_effect)
        self._banner_anim = QPropertyAnimation(self._banner_effect, b"opacity", self)
        self._banner_anim.setDuration(600)
        self._banner_anim.setEasingCurve(QEasingCurve.OutQuad)


        # item mapping: dot_id -> dot reference
        self.dot_index: List[Tuple[int, int]] = []  # list of (page_index, dot_list_index)
        self.dot_ellipse_items: Dict[int, DotEllipseItem] = {}
        self.dot_label_items: Dict[int, LabelTextItem] = {}
        self.dot_connector_items: Dict[int, QGraphicsLineItem] = {}


        # Model highlight state (for totals table selection -> halo overlays)
        self.active_highlight_model: Optional[str] = None
        self.halo_items: List[QGraphicsEllipseItem] = []
        self.model_colors: Dict[str, QColor] = {}  # model -> QColor (derived from sampled HSV)

        # Label sizing (slider controls)
        # Slider range is -100..+200 (percent delta from base).
        self.label_scale_pct = 0
        self.base_label_font_pt = 14  # must match LabelTextItem default
        self.base_export_pdf_font_pt = 8
        self.base_export_img_font_scale = 0.55

        # Halo breathing animation
        self.halo_anim_timer = QTimer(self)
        self.halo_anim_timer.setInterval(30)  # ~33 FPS
        self.halo_anim_timer.timeout.connect(self._tick_halo_breathe)

        self.halo_anim_clock = QElapsedTimer()
        self.halo_breathe_period_ms = 1000  # 1 second for 100->0->100 loop

        # Mode flags
        self.mode_pick_sample = False
        self.mode_add_dot = False
        self.mode_remove_dot = False

        # Naming/text-selection mode (drag rectangle to capture PDF words)
        self.mode_text_select = False
        self.text_select_drag_start_scene: Optional[QPointF] = None
        self.text_select_rect_item = None
        self.text_select_callback: Optional[Callable[[str], None]] = None

        # Naming workflow “resume” cursor
        self.naming_last_uid: Optional[str] = None

        # While naming: keep NAMES tab selected + highlight/scroll active device row
        self.force_names_tab_selected: bool = False
        self.names_highlight_uid: Optional[str] = None

        # Single-dot focus halo for the naming wizard (separate from model halos)
        self.focus_dot_uid: Optional[str] = None
        self.focus_halo_items: List[QGraphicsEllipseItem] = []

        # Mouse UX
        self.last_hover_scene_pos: Optional[QPointF] = None

        # Magnifier popup (for pixel-accurate sampling) — floating tooltip windows (never clipped)
        self.magnifier = QLabel()
        self.magnifier.setFixedSize(180, 180)
        self.magnifier.setWindowFlags(Qt.ToolTip | Qt.FramelessWindowHint)
        self.magnifier.setAttribute(Qt.WA_TransparentForMouseEvents, True)
        self.magnifier.setStyleSheet("background: white; border: 2px solid black;")
        self.magnifier.hide()

        self.magnifier_info = QLabel()
        self.magnifier_info.setWindowFlags(Qt.ToolTip | Qt.FramelessWindowHint)
        self.magnifier_info.setAttribute(Qt.WA_TransparentForMouseEvents, True)
        self.magnifier_info.setStyleSheet("background: white; border: 1px solid black; padding: 3px;")
        self.magnifier_info.hide()
        # ---------------------------
        # Magnifier performance: cache + throttle
        # ---------------------------
        self._page_bgr_cache: Dict[Tuple[int, float], np.ndarray] = {}
        self._page_hsv_cache: Dict[Tuple[int, float], np.ndarray] = {}
        self._mag_pending_scene_pos: Optional[QPointF] = None
        self._mag_timer = QTimer(self)
        self._mag_timer.setInterval(16)  # ~60fps cap
        self._mag_timer.timeout.connect(self._magnifier_tick)

        # Auto-hide magnifier when app loses focus (prevents it showing on top of other apps)
        QApplication.instance().applicationStateChanged.connect(self._on_app_state_changed)

        # Drag state (dot dragging + label move finalize)
        self.dragging_dot_id: Optional[int] = None
        self.drag_dot_offset_scene = QPointF(0, 0)
        self.is_dragging_dot = False

        # Middle-mouse pan state
        self.is_mid_panning = False
        self.mid_pan_start_pos = None  # QPoint
        self.mid_pan_start_h = 0
        self.mid_pan_start_v = 0

        # Detection tuning (can later move to UI sliders)
        self.detect_area_min = 28
        self.detect_area_max = 8000
        self.detect_circ_min = 0.35
        # ---------------------------
        # Feature 1: Dot size tri-knob (min / expected / max) + marching-ants previews
        # Stored in PDF units (points) so it scales with render_zoom.
        # ---------------------------
        self.dot_r_min_pdf = 4.0       # min dot radius (PDF points)
        self.dot_r_expected_pdf = 7.0  # expected dot radius (PDF points)
        self.dot_r_max_pdf = 14.0       # max dot radius (PDF points)
        self._dot_expected_last = int(round(self.dot_r_expected_pdf))

        # 3 preview circles (min/expected/max) that follow hover/cursor in scene coords
        self._dot_size_preview_items: List[QGraphicsEllipseItem] = []
        self._dot_size_preview_visible = False

        # Marching-ants animation
        self._ants_timer = QTimer(self)
        self._ants_timer.setInterval(40)  # ~25fps
        self._ants_timer.timeout.connect(self._tick_marching_ants)
        self._ants_phase = 0.0


        # Hue tolerance (tightened slightly)
        self.detect_dh = 10

        # IMPORTANT: allow pastel models like C3 to detect
        self.detect_min_s = 18
        self.detect_min_v = 35

        # Dedupe distance for same-dot double hits
        self.dedupe_dist_px = 10

        # Extra gating to prevent cross-model confusion (ex: C1 vs C14)
        self.detect_max_hue_dist = 7
        self.detect_max_sat_diff = 60
        self.detect_max_val_diff = 70

        # Build UI
        self._build_ui()
        # ESC always exits tool modes (reliable regardless of focus)
        self.showMaximized()
        self.esc_shortcut = QShortcut(QKeySequence(Qt.Key_Escape), self)
        self.esc_shortcut.activated.connect(lambda: self.exit_tool_modes("Exited tool mode (ESC)."))

    # ---------------- UI ----------------

    def _build_ui(self):
        root = QWidget()
        self.setCentralWidget(root)

        splitter = QSplitter()
        layout = QHBoxLayout(root)
        layout.addWidget(splitter)

        # Left panel
        left_panel = QWidget()
        left_layout = QVBoxLayout(left_panel)

        btn_open = QPushButton("Open PDF (Ctrl+Space bar)")
        btn_open.clicked.connect(self.open_pdf)
        left_layout.addWidget(btn_open)

        self.short_open_pdf = QShortcut(QKeySequence("Ctrl+Space"), self)
        self.short_open_pdf.setContext(Qt.ApplicationShortcut)
        self.short_open_pdf.activated.connect(self.open_pdf)

        self.status_label = QLabel(" ")
        self.status_label.setWordWrap(True)
        left_layout.addWidget(self.status_label)

        zoom_box = QGroupBox("Zoom")
        zb = QHBoxLayout(zoom_box)
        btn_z_in = QPushButton("+")
        btn_z_out = QPushButton("-")
        btn_z_in.clicked.connect(lambda: self.view.scale(1.2, 1.2))
        btn_z_out.clicked.connect(lambda: self.view.scale(1 / 1.2, 1 / 1.2))

        sample_box = QGroupBox("Sampling & Detection")
        sb = QVBoxLayout(sample_box)

        btn_pick = QPushButton("Pick Color Sample (Ctrl+W)")
        btn_pick.clicked.connect(self.enable_pick_sample_mode)
        sb.addWidget(btn_pick)

        self.short_pick_sample = QShortcut(QKeySequence("Ctrl+W"), self)
        self.short_pick_sample.setContext(Qt.ApplicationShortcut)
        self.short_pick_sample.activated.connect(self.enable_pick_sample_mode)

        btn_preview = QPushButton("Generate Preview (All Pages)")
        btn_preview.clicked.connect(self.generate_preview_all_pages)
        sb.addWidget(btn_preview)

        btn_preview_page = QPushButton("Generate Preview (Current Page Only)")
        btn_preview_page.clicked.connect(self.generate_preview_current_page)
        sb.addWidget(btn_preview_page)

        left_layout.addWidget(sample_box)

        edit_box = QGroupBox("Edit")
        eb = QVBoxLayout(edit_box)

        btn_add = QPushButton("Add Device")
        btn_add.clicked.connect(self.enable_add_dot_mode)
        eb.addWidget(btn_add)

        btn_remove = QPushButton("Edit Device")
        btn_remove.clicked.connect(self.enable_remove_dot_mode)
        eb.addWidget(btn_remove)

        btn_clear = QPushButton("Clear Devices")
        btn_clear.clicked.connect(self.clear_current_page_dots)
        eb.addWidget(btn_clear)

        self.btn_exclude = QPushButton("Draw Exclusion Zone (Ctrl+Q)")
        self.btn_exclude.clicked.connect(self.enable_exclusion_zone_mode)
        eb.addWidget(self.btn_exclude)

        self.short_exclude_zone = QShortcut(QKeySequence("Ctrl+Q"), self)
        self.short_exclude_zone.setContext(Qt.ApplicationShortcut)
        self.short_exclude_zone.activated.connect(self.enable_exclusion_zone_mode)

        btn_clear_excl = QPushButton("Clear Exclusion Zones")
        btn_clear_excl.clicked.connect(self.clear_current_page_exclusions)
        eb.addWidget(btn_clear_excl)

        left_layout.addWidget(edit_box)

        export_box = QGroupBox("Export")
        xb = QVBoxLayout(export_box)

        btn_export_pdf = QPushButton("Export Labeled PDF (vector)")
        btn_export_pdf.clicked.connect(self.export_labeled_pdf)
        xb.addWidget(btn_export_pdf)

        btn_export_imgs = QPushButton("Export Images (choose DPI)")
        btn_export_imgs.clicked.connect(self.export_images)
        xb.addWidget(btn_export_imgs)

        # Label Size slider (-100% .. +200%)
        self.lbl_label_scale = QLabel("Label Size: 100%")
        xb.addWidget(self.lbl_label_scale)

        self.slider_label_scale = QSlider(Qt.Horizontal)
        self.slider_label_scale.setRange(-100, 200)
        self.slider_label_scale.setValue(0)  # 0% delta => 100%
        self.slider_label_scale.setTickInterval(25)
        self.slider_label_scale.setSingleStep(5)
        self.slider_label_scale.valueChanged.connect(self.on_label_scale_changed)
        xb.addWidget(self.slider_label_scale)

        left_layout.addWidget(export_box)
        # ---------------------------
        # Feature 1: Dot Size Controls (tri "knob" via 3 sliders)
        # ---------------------------
        dot_size_box = QGroupBox("Dot Size (min / expected / max)")
        ds = QVBoxLayout(dot_size_box)

        self.lbl_dot_size = QLabel(
            f"Min: {self.dot_r_min_pdf:.1f}  Expected: {self.dot_r_expected_pdf:.1f}  Max: {self.dot_r_max_pdf:.1f} (PDF pts)"
        )
        ds.addWidget(self.lbl_dot_size)
        # NEW: marching-ants preview widget under Dot Size controls
        self.dot_size_preview_widget = DotSizePreviewWidget()
        self.dot_size_preview_widget.set_radii(self.dot_r_min_pdf, self.dot_r_expected_pdf, self.dot_r_max_pdf)
        ds.addWidget(self.dot_size_preview_widget)


        def _mk_slider():
            s = QSlider(Qt.Horizontal)
            s.setRange(1, 20)        # PDF points range (tune later)
            s.setTickInterval(1)
            s.setSingleStep(1)
            return s

        self.slider_dot_min = _mk_slider()
        self.slider_dot_expected = _mk_slider()
        self.slider_dot_max = _mk_slider()

        self.slider_dot_min.setValue(int(round(self.dot_r_min_pdf)))
        self.slider_dot_expected.setValue(int(round(self.dot_r_expected_pdf)))
        self.slider_dot_max.setValue(int(round(self.dot_r_max_pdf)))

        ds.addWidget(QLabel("Min radius"))
        ds.addWidget(self.slider_dot_min)
        ds.addWidget(QLabel("Expected radius"))
        ds.addWidget(self.slider_dot_expected)
        ds.addWidget(QLabel("Max radius"))
        ds.addWidget(self.slider_dot_max)

        self.slider_dot_min.valueChanged.connect(self.on_dot_size_sliders_changed)
        self.slider_dot_expected.valueChanged.connect(self.on_dot_size_sliders_changed)
        self.slider_dot_max.valueChanged.connect(self.on_dot_size_sliders_changed)

        left_layout.addWidget(dot_size_box)

        left_layout.addStretch(1)

        # Center viewer
        viewer_panel = QWidget()
        viewer_layout = QVBoxLayout(viewer_panel)

        viewer_layout.addWidget(self.view)

        # Right panel: Samples + Ordering
        # Right panel: Totals (top) + Model Order + Collapsible Samples (bottom)
        right_panel = QWidget()
        right_layout = QVBoxLayout(right_panel)

        # ---------------------------
        # TOP: Page List + Totals Tabs
        # ---------------------------
        right_layout.addWidget(QLabel("Floors (Pages) + Per-Model Totals"))

        # Page list (for delete/rename floor titles)
        self.page_list = QListWidget()
        self.page_list.setSelectionMode(QAbstractItemView.SingleSelection)
        self.page_list.setMaximumHeight(130)
        right_layout.addWidget(self.page_list)
        self.page_list.itemClicked.connect(self.on_floor_clicked)


        page_btn_row = QWidget()
        page_btn_row_l = QHBoxLayout(page_btn_row)
        page_btn_row_l.setContentsMargins(0, 0, 0, 0)

        self.btn_rename_floor = QPushButton("Rename Floor")
        self.btn_delete_floor = QPushButton("Delete Floor (remove from program)")

        self.btn_rename_floor.clicked.connect(self.rename_selected_floor)
        self.btn_delete_floor.clicked.connect(self.delete_selected_floor)

        page_btn_row_l.addWidget(self.btn_rename_floor)
        page_btn_row_l.addWidget(self.btn_delete_floor)
        right_layout.addWidget(page_btn_row)

        # Totals tabs (one tab per page that contains bubbles)
        self.totals_tabs = QTabWidget()
        self.totals_tabs.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        # Double-click tab title to rename
        self.totals_tabs.tabBarDoubleClicked.connect(self.rename_floor_tab)

        right_layout.addWidget(self.totals_tabs, 3)

        name_row = QWidget()
        name_row_l = QHBoxLayout(name_row)
        name_row_l.setContentsMargins(0, 0, 0, 0)

        btn_name_devices = QPushButton("Name Devices…")
        btn_name_devices.clicked.connect(self.open_device_naming_window)

        self.btn_delete_model = QPushButton("Delete device type")
        self.btn_delete_model.clicked.connect(self.delete_selected_device_type)

        name_row_l.addWidget(btn_name_devices)
        name_row_l.addWidget(self.btn_delete_model)
        right_layout.addWidget(name_row)

        # NEW: Export totals + devices tables
        self.btn_export_table = QPushButton("Export Table…")
        self.btn_export_table.clicked.connect(self.export_totals_and_devices_tables)
        right_layout.addWidget(self.btn_export_table)

        # ---------------------------
        # BOTTOM: Collapsible Samples table (MINIMIZED by default)
        # ---------------------------
        self.samples_group = QGroupBox("Model Samples (collapsed by default)")
        self.samples_group.setCheckable(True)
        self.samples_group.setChecked(False)  # minimized by default

        sg_layout = QVBoxLayout(self.samples_group)

        self.sample_table = QTableWidget(0, 2)
        self.sample_table.setHorizontalHeaderLabels(["Model", "HSV"])
        self.sample_table.horizontalHeader().setStretchLastSection(True)
        self.sample_table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.sample_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.sample_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.sample_table.verticalHeader().setVisible(False)

        sg_layout.addWidget(self.sample_table)

        right_layout.addWidget(self.samples_group, 1)

        splitter.addWidget(left_panel)
        splitter.addWidget(viewer_panel)
        splitter.addWidget(right_panel)
        splitter.setStretchFactor(0, 0)
        splitter.setStretchFactor(1, 1)
        splitter.setStretchFactor(2, 0)

        self.resize(1650, 950)

    # ---------------- PDF Loading / Rendering ----------------

    def show_viewport_banner(self, text: str, ms: int = 3000):
        """Show a transient banner at the top of the viewport, then fade it out."""
        if not hasattr(self, "_banner"):
            return

        # Stop any existing animation and show fully opaque
        try:
            self._banner_anim.stop()
        except Exception:
            pass
        try:
            self._banner_effect.setOpacity(1.0)
        except Exception:
            pass

        self._banner.setText(text)

        vp = self.view.viewport()
        max_w = max(200, vp.width() - 20)
        self._banner.setMaximumWidth(max_w)
        self._banner.adjustSize()
        self._banner.move(10, 10)
        self._banner.show()

        QTimer.singleShot(ms, self._fade_out_viewport_banner)

    def _fade_out_viewport_banner(self):
        if not hasattr(self, "_banner") or not self._banner.isVisible():
            return

        # Avoid stacking finished connections
        try:
            self._banner_anim.finished.disconnect()
        except Exception:
            pass
        self._banner_anim.finished.connect(self._banner.hide)

        self._banner_anim.stop()
        self._banner_anim.setStartValue(1.0)
        self._banner_anim.setEndValue(0.0)
        self._banner_anim.start()

    def on_floor_clicked(self, item: QListWidgetItem):
        doc_pi = item.data(Qt.UserRole)
        if doc_pi is None:
            return
        self.focus_page(doc_pi)

    def _invalidate_page_cache(self):
        self._page_bgr_cache.clear()
        self._page_hsv_cache.clear()

    def _magnifier_tick(self):
        if not self._mag_pending_scene_pos:
            self._mag_timer.stop()
            return
        pos = self._mag_pending_scene_pos
        self._mag_pending_scene_pos = None
        self.update_magnifier(pos)


    def _darker_qcolor(self, c: QColor, delta: int = 20) -> QColor:
        """
        Return QColor darkened by subtracting delta from RGB channels (clamped).
        Keeps alpha the same.
        """
        r = max(0, c.red()   - delta)
        g = max(0, c.green() - delta)
        b = max(0, c.blue()  - delta)
        return QColor(r, g, b, c.alpha())

    # ---------------------------
    # Feature 1: Dot size tri sliders + marching-ants preview circles
    # ---------------------------

    def on_dot_size_sliders_changed(self, _value: int):
        """
        Enforce: min <= expected <= max

        Behavior:
        - If EXPECTED moves, shift MIN and MAX by the same delta (clamped to slider range),
          so the min/expected/max window moves together.
        - If MIN or MAX moves, just enforce ordering as before.
        """
        sender = self.sender()

        mn = int(self.slider_dot_min.value())
        ex = int(self.slider_dot_expected.value())
        mx = int(self.slider_dot_max.value())

        # ---------------------------------------------------------
        # NEW: If expected changed, shift min & max along with it
        # ---------------------------------------------------------
        if sender is self.slider_dot_expected:
            prev = getattr(self, "_dot_expected_last", ex)
            delta = ex - prev

            if delta != 0:
                smin = int(self.slider_dot_min.minimum())
                smax = int(self.slider_dot_max.maximum())

                # Clamp delta so (mn+delta) stays >= smin AND (mx+delta) stays <= smax
                if mn + delta < smin:
                    delta = smin - mn
                if mx + delta > smax:
                    delta = smax - mx

                if delta != 0:
                    new_mn = mn + delta
                    new_mx = mx + delta

                    # Apply without recursive signal storms
                    self.slider_dot_min.blockSignals(True)
                    self.slider_dot_max.blockSignals(True)
                    self.slider_dot_min.setValue(new_mn)
                    self.slider_dot_max.setValue(new_mx)
                    self.slider_dot_min.blockSignals(False)
                    self.slider_dot_max.blockSignals(False)

                    mn = new_mn
                    mx = new_mx

        # Clamp ordering (same as before)
        if mn > ex:
            ex = mn
            self.slider_dot_expected.blockSignals(True)
            self.slider_dot_expected.setValue(ex)
            self.slider_dot_expected.blockSignals(False)

        if ex > mx:
            mx = ex
            self.slider_dot_max.blockSignals(True)
            self.slider_dot_max.setValue(mx)
            self.slider_dot_max.blockSignals(False)

        # Keep expected tracking updated AFTER any ordering clamps
        self._dot_expected_last = int(ex)

        # Store as PDF points
        self.dot_r_min_pdf = float(mn)
        self.dot_r_expected_pdf = float(ex)
        self.dot_r_max_pdf = float(mx)

        self.lbl_dot_size.setText(
            f"Min: {self.dot_r_min_pdf:.1f}  Expected: {self.dot_r_expected_pdf:.1f}  Max: {self.dot_r_max_pdf:.1f} (PDF pts)"
        )

        # Keep the DotSizePreviewWidget synced
        if hasattr(self, "dot_size_preview_widget") and self.dot_size_preview_widget:
            self.dot_size_preview_widget.set_radii(self.dot_r_min_pdf, self.dot_r_expected_pdf, self.dot_r_max_pdf)

        # Update preview circles in-scene if they are visible
        if getattr(self, "_dot_size_preview_visible", False):
            self.update_dot_size_previews()

    def _ensure_dot_size_preview_items(self):
        """
        Create the 3 scene ellipse items once.
        """
        if self._dot_size_preview_items:
            return

        for _ in range(3):
            it = QGraphicsEllipseItem()
            it.setBrush(Qt.NoBrush)
            it.setZValue(200)  # above labels/halos
            self.scene.addItem(it)
            self._dot_size_preview_items.append(it)

    def set_dot_size_preview_visible(self, visible: bool):
        """
        Show/hide the min/expected/max preview circles + start/stop ants timer.
        """
        self._dot_size_preview_visible = bool(visible)

        if not visible:
            for it in self._dot_size_preview_items:
                it.setVisible(False)

            # Stop ants ONLY if the widget isn't visible either
            widget_visible = bool(
                getattr(self, "dot_size_preview_widget", None)) and self.dot_size_preview_widget.isVisible()
            if not widget_visible and self._ants_timer.isActive():
                self._ants_timer.stop()
            return

        self._ensure_dot_size_preview_items()
        for it in self._dot_size_preview_items:
            it.setVisible(True)

        if not self._ants_timer.isActive():
            self._ants_timer.start()

        self.update_dot_size_previews()

    def update_dot_size_previews(self):
        """
        Position 3 circles (min/expected/max) around the current hover point.
        Radii are in PDF points -> converted to px using render_zoom, so they scale with zoom.
        """
        if not self._dot_size_preview_visible:
            return
        if not self.last_hover_scene_pos:
            return
        if not self.doc:
            return

        self._ensure_dot_size_preview_items()

        # Scene position -> page-local px so we can compute circles in scene coords correctly
        page_index, x_px, y_px = self.scene_to_page_px(self.last_hover_scene_pos)
        if page_index is None:
            for it in self._dot_size_preview_items:
                it.setVisible(False)
            return

        # Convert PDF radii -> px
        r_min_px = self.dot_r_min_pdf * self.render_zoom
        r_exp_px = self.dot_r_expected_pdf * self.render_zoom
        r_max_px = self.dot_r_max_pdf * self.render_zoom

        # Center in scene coordinates (we want the ring centered on the hover point)
        center_scene = self.page_px_to_scene(page_index, x_px, y_px)

        radii = [r_min_px, r_exp_px, r_max_px]

        # Styles: dashed "marching ants"
        # - min: thin
        # - expected: medium
        # - max: thicker
        widths = [1, 2, 2]

        for it, r_px, w in zip(self._dot_size_preview_items, radii, widths):
            it.setRect(center_scene.x() - r_px, center_scene.y() - r_px, r_px * 2, r_px * 2)

            pen = QPen(QColor(0, 0, 0), w)
            pen.setStyle(Qt.DashLine)
            pen.setDashPattern([4, 4])
            pen.setDashOffset(self._ants_phase)
            it.setPen(pen)
            it.setVisible(True)

    def _tick_marching_ants(self):
        """
        Advances dash offset for the preview circles.
        """
        # Keep ants running if either:
        # - scene hover preview circles are visible, OR
        # - the DotSizePreviewWidget is visible
        widget_visible = bool(
            getattr(self, "dot_size_preview_widget", None)) and self.dot_size_preview_widget.isVisible()
        if not self._dot_size_preview_visible and not widget_visible:
            if self._ants_timer.isActive():
                self._ants_timer.stop()
            return

        self._ants_phase += 1.0
        if self._ants_phase > 10000:
            self._ants_phase = 0.0

        # Only update pen dash offset (cheap)
        for it in self._dot_size_preview_items:
            pen = it.pen()
            pen.setDashOffset(self._ants_phase)
            it.setPen(pen)
        # NEW: animate the widget too
        if hasattr(self, "dot_size_preview_widget") and self.dot_size_preview_widget:
            self.dot_size_preview_widget.set_dash_offset(self._ants_phase)



    def focus_page(self, doc_page_index: int):
        """
        Scroll to a page and zoom so it fits the view.
        """
        if doc_page_index not in self.doc_to_visible_index:
            return

        vis_i = self.doc_to_visible_index[doc_page_index]
        y0 = self.page_y_offsets[vis_i]
        w, h = self.page_sizes_px[vis_i]

        rect = QRectF(0, y0, w, h)

        # Zoom out to fit the whole page
        self.view.fitInView(rect, Qt.KeepAspectRatio)

        # Center on the page (nice navigation feel)
        self.view.centerOn(rect.center())
        # Feature 1: update previews after zoom/fit changes
        self.update_dot_size_previews()


    def open_device_naming_window(self):
        dots = self.get_all_dots_sorted_by_number()
        if not dots:
            QMessageBox.information(self, "Device Naming", "No detected/labeled dots exist yet.")
            return

        uids = [d.uid for d in dots]

        # Force NAMES tab to stay selected during naming
        self.force_names_tab_selected = True
        self.highlight_names_uid(uids[0])  # initialize highlight/scroll

        # If dialog already exists and is open, just bring it forward
        if hasattr(self, "device_naming_dialog") and self.device_naming_dialog is not None:
            try:
                if self.device_naming_dialog.isVisible():
                    self.device_naming_dialog.raise_()
                    self.device_naming_dialog.activateWindow()
                    return
            except Exception:
                pass

        # IMPORTANT: keep a reference so it doesn't get garbage-collected
        self.device_naming_dialog = DeviceNamingDialog(self, uids)

        # Non-modal: lets you click/drag on the PDF while dialog is open
        self.device_naming_dialog.setModal(False)

        # When the dialog closes: stop forcing NAMES, and clear the single-dot focus halo
        def _on_naming_closed(_result):
            self.force_names_tab_selected = False
            self.names_highlight_uid = None  # stop auto-selection behavior
            self.clear_focus_halo()  # stop the "current dot" halo
            self._stop_halo_timer_if_idle()  # stop timer if no halos remain
            self.refresh_totals_ui()

        self.device_naming_dialog.finished.connect(_on_naming_closed)

        self.device_naming_dialog.show()
        self.device_naming_dialog.raise_()
        self.device_naming_dialog.activateWindow()

    def open_pdf(self):
        dlg = PdfOpenDialog(self)
        dlg.setWindowTitle("Select PDF")

        if dlg.exec() != QDialog.Accepted:
            return

        files = dlg.selectedFiles()
        path = files[0] if files else ""
        if not path:
            return


        try:
            self.doc = fitz.open(path)
            self._invalidate_page_cache()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to open PDF:\n{e}")
            return

        self.pdf_path = path
        self.samples.clear()
        self.model_order.clear()
        self.dots_by_page.clear()
        self.exclusion_zones_by_page.clear()  # <-- add this

        self.refresh_samples_ui()
        self.refresh_model_order_ui()

        self.removed_pages.clear()
        self.page_titles.clear()

        self.render_all_pages()
        self.show_viewport_banner(f"Loaded: {path}   (Pages: {len(self.doc)})", ms=3000)

        self.refresh_pages_list_ui()
        self.refresh_totals_ui()

    def delete_selected_device_type(self):
        """
        Deletes a model/device type from:
          - model_order
          - samples (optional)
          - all dots on all visible pages (removes circles/labels)
        Uses the currently-selected row in the active totals table if possible,
        otherwise prompts.
        """
        models = self.get_model_list()
        if not models:
            QMessageBox.information(self, "Delete device type", "No device types/models exist yet.")
            return

        selected_model = None

        # Try to infer from current tab selection (TotalsTabWidget or QTableWidget)
        cur = self.totals_tabs.currentWidget()
        table = None
        if isinstance(cur, TotalsTabWidget):
            table = cur.main_table
        elif isinstance(cur, QTableWidget):
            table = cur

        if table and table.currentRow() >= 0:
            it = table.item(table.currentRow(), 0)
            if it:
                selected_model = it.text().strip()

        # Fallback: prompt
        if not selected_model or selected_model not in models:
            selected_model, ok = QInputDialog.getItem(
                self,
                "Delete device type",
                "Select model/device type to delete:",
                models,
                0,
                False
            )
            if not ok or not selected_model:
                return

        model = selected_model.strip()

        confirm = QMessageBox.question(
            self,
            "Delete device type",
            f"Delete model '{model}'?\n\n"
            f"- Removes it from the model list/order\n"
            f"- Removes its sample (if present)\n"
            f"- Deletes ALL dots of this model from all pages\n\n"
            f"This cannot be undone.",
            QMessageBox.Yes | QMessageBox.No
        )
        if confirm != QMessageBox.Yes:
            return

        # Remove dots of this model across all visible pages
        for pi in list(self.visible_pages):
            dots = self.dots_by_page.get(pi, [])
            if dots:
                self.dots_by_page[pi] = [d for d in dots if d.model != model]

        # Remove from ordering + samples + cached colors
        if model in self.model_order:
            self.model_order = [m for m in self.model_order if m != model]
        if model in self.samples:
            self.samples.pop(model, None)
        if model in self.model_colors:
            self.model_colors.pop(model, None)

        # Clear highlight if it was on this model
        if self.active_highlight_model == model:
            self.clear_model_highlight()

        # Rebuild everything
        self.refresh_samples_ui()
        self.renumber_all_labels()
        self.rebuild_overlay_items()
        self.refresh_pages_list_ui()
        self.refresh_totals_ui()

        # Keep halos consistent if some other model is highlighted
        if self.active_highlight_model:
            self.apply_model_halos(self.active_highlight_model)

    def change_render_zoom(self, factor: float):
        if not self.doc:
            return
        self.render_zoom = max(0.75, min(6.0, self.render_zoom * factor))
        self.render_all_pages()

    def render_all_pages(self):
        """
        Continuous scroll: stack visible pages vertically in the scene.
        (Pages in self.removed_pages are hidden from the program.)
        """
        if not self.doc:
            return
        self._invalidate_page_cache()

        self.scene.clear()
        self.page_y_offsets.clear()
        self.page_sizes_px.clear()
        self.page_pixmaps.clear()

        # Build visible page list (exclude removed pages)
        self.visible_pages = [i for i in range(len(self.doc)) if i not in self.removed_pages]
        self.doc_to_visible_index = {doc_i: vis_i for vis_i, doc_i in enumerate(self.visible_pages)}

        y_cursor = 0.0
        page_gap = 30  # pixels

        for doc_page_index in self.visible_pages:
            page = self.doc[doc_page_index]
            mat = fitz.Matrix(self.render_zoom, self.render_zoom)
            pix = page.get_pixmap(matrix=mat, alpha=False)

            rgb = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.height, pix.width, 3)
            qimg = QImage(rgb.data, pix.width, pix.height, pix.stride, QImage.Format_RGB888)
            pixmap = QPixmap.fromImage(qimg.copy())

            item = QGraphicsPixmapItem(pixmap)
            item.setPos(0, y_cursor)
            self.scene.addItem(item)

            self.page_y_offsets.append(y_cursor)
            self.page_sizes_px.append((pix.width, pix.height))
            self.page_pixmaps.append(item)

            # draw a subtle border around each page
            border = self.scene.addRect(QRectF(0, y_cursor, pix.width, pix.height), QPen(QColor(180, 180, 180), 1))
            border.setZValue(1)

            y_cursor += pix.height + page_gap

        self.scene.setSceneRect(self.scene.itemsBoundingRect())

        # Redraw overlays after re-render
        self.rebuild_overlay_items()

    # ---------------- Modes ----------------

    def enable_pick_sample_mode(self):
        self.exit_tool_modes()  # <-- ensures exclusion zone + other modes are OFF
        self.mode_pick_sample = True

        # Make sampling precise (no hand cursor)
        self.view.setDragMode(QGraphicsView.NoDrag)
        self.view.viewport().setCursor(Qt.CrossCursor)

        self.magnifier.setVisible(True)
        self.magnifier_info.setVisible(True)

        self.statusBar().showMessage(
            "Pick Sample mode: hover to preview pixels; click a dot; enter model (C11, C1, etc).")

    def enable_add_dot_mode(self):
        self.exit_tool_modes()  # <-- ensures exclusion zone + other modes are OFF
        self.mode_add_dot = True

        self.view.setDragMode(QGraphicsView.NoDrag)
        self.view.viewport().setCursor(Qt.CrossCursor)
        self.magnifier.setVisible(False)
        self.magnifier_info.setVisible(False)
        self.statusBar().showMessage("Add Dot mode: click location to add; model will be predicted (you can override).")

    def enable_remove_dot_mode(self):
        self.exit_tool_modes()  # <-- ensures exclusion zone + other modes are OFF
        self.mode_remove_dot = True

        self.view.setDragMode(QGraphicsView.NoDrag)
        self.view.viewport().setCursor(Qt.CrossCursor)
        self.magnifier.setVisible(False)
        self.magnifier_info.setVisible(False)
        self.statusBar().showMessage("Remove Dot mode: click a dot to exclude/delete via dialog.")

    def enable_exclusion_zone_mode(self):
        self.exit_tool_modes()  # <-- ensures pick/add/remove are OFF
        self.mode_exclusion_zone = True

        self.view.setDragMode(QGraphicsView.NoDrag)
        self.view.viewport().setCursor(Qt.CrossCursor)

        self.statusBar().showMessage("Exclusion Zone mode: click-drag to draw a rectangle to ignore during detection.")

    def clear_current_page_exclusions(self):
        if not self.doc:
            return
        center = self.view.mapToScene(self.view.viewport().rect().center())
        page_index, _, _ = self.scene_to_page_px(center)
        if page_index is None:
            return
        self.exclusion_zones_by_page[page_index] = []
        self.rebuild_overlay_items()
        self.statusBar().showMessage("Cleared exclusion zones for current page.")

    def exit_tool_modes(self, msg: str = "Default mode active."):
        # turn off modes
        self.mode_pick_sample = False
        self.mode_add_dot = False
        self.mode_remove_dot = False
        self.mode_exclusion_zone = False
        self.mode_text_select = False
        self.text_select_callback = None

        # reset drag states
        self.dragging_dot_id = None
        self.is_dragging_dot = False
        self.exclusion_drag_start_scene = None

        # restore normal navigation
        self.view.setDragMode(QGraphicsView.ScrollHandDrag)
        self.view.viewport().unsetCursor()

        # hide magnifier
        self.magnifier.setVisible(False)
        self.magnifier_info.setVisible(False)
        # Feature 1: hide dot-size previews
        if self._dot_size_preview_visible:
            self.set_dot_size_preview_visible(False)


        self.statusBar().showMessage(msg)

    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Escape:
            self.exit_tool_modes("Exited tool mode (ESC).")
            event.accept()
            return
        super().keyPressEvent(event)

    def _on_app_state_changed(self, state):
        """
        If user alt-tabs away, hide the magnifier so it doesn't float over other programs.
        If user returns and Pick Sample mode is active, show it again.
        """
        if state != Qt.ApplicationActive:
            self.magnifier.hide()
            self.magnifier_info.hide()
            return

        # App is active again
        if self.mode_pick_sample and self.last_hover_scene_pos is not None:
            # Optional: refresh once when returning focus
            self.update_magnifier(self.last_hover_scene_pos)

    # ---------------- Input Handling ----------------

    def eventFilter(self, source, event):
        if source != self.view.viewport():
            return super().eventFilter(source, event)
        # ---------------------------
        # ESC = EXIT ACTIVE TOOL MODE
        # ---------------------------
        if event.type() == QEvent.Type.KeyPress:
            # ESC exits tool modes
            if event.key() == Qt.Key_Escape:
                self.exit_tool_modes("Exited tool mode (ESC).")
                return True

            # DEL deletes selected dot/label and renumbers
            if event.key() == Qt.Key_Delete:
                selected = self.scene.selectedItems()

                # Prefer deleting by selected label (LabelTextItem)
                for it in selected:
                    if isinstance(it, LabelTextItem):
                        dot_id = it.dot_id
                        # clear selection first so we don't keep a stale selected item
                        it.setSelected(False)

                        if self.delete_dot_by_id(dot_id):
                            self.statusBar().showMessage("Deleted dot. Renumbered.")
                        return True

                # Also allow deleting by selecting the dot ring itself
                for it in selected:
                    if isinstance(it, DotEllipseItem):
                        dot_id = it.dot_id
                        it.setSelected(False)

                        if self.delete_dot_by_id(dot_id):
                            self.statusBar().showMessage("Deleted dot. Renumbered.")
                        return True

                return True

        # ---------------------------
        # CTRL + WHEEL = ZOOM (keep normal wheel scroll otherwise)
        # ---------------------------
        # ---------------------------
        # CTRL + WHEEL = ZOOM (keep normal wheel scroll otherwise)
        # SHIFT + WHEEL = HORIZONTAL SCROLL (MS-style)
        # ---------------------------
        if event.type() == QEvent.Type.Wheel:
            wheel: QWheelEvent = event

            # CTRL + wheel = zoom
            if wheel.modifiers() & Qt.ControlModifier:
                delta = wheel.angleDelta().y()
                factor = 1.15 if delta > 0 else (1 / 1.15)
                self.view.scale(factor, factor)
                # Feature 1: keep previews in sync with zoom changes
                self.update_dot_size_previews()
                return True

            # ALT + wheel = adjust Expected dot radius slider
            # (On some systems Alt+wheel is delivered as horizontal wheel => angleDelta().x(), and
            #  sometimes Alt isn't present in wheel.modifiers(). So we check both and accept X or Y.)
            mods = wheel.modifiers() | QApplication.keyboardModifiers()
            if mods & Qt.AltModifier:
                ad = wheel.angleDelta()
                pd = wheel.pixelDelta()

                # Prefer angleDelta (notches). If it's 0 on Y, use X (horizontal wheel).
                raw = ad.y() if ad.y() != 0 else ad.x()

                # Trackpads may use pixelDelta instead
                if raw == 0:
                    raw = pd.y() if pd.y() != 0 else pd.x()

                if raw == 0:
                    return True  # consume so it doesn't horizontal-scroll the view

                # 1 notch = 1 unit; Shift+Alt = faster
                step = 5 if (mods & Qt.ShiftModifier) else 1
                direction = 1 if raw > 0 else -1

                cur = self.slider_dot_expected.value()
                new = cur + (direction * step)
                new = max(self.slider_dot_expected.minimum(),
                          min(self.slider_dot_expected.maximum(), new))

                if new != cur:
                    self.slider_dot_expected.setValue(new)

                return True

            # SHIFT + wheel = horizontal scroll
            if wheel.modifiers() & Qt.ShiftModifier:
                delta = wheel.angleDelta().y()

                # Shift+Scroll UP => scroll LEFT
                # Shift+Scroll DOWN => scroll RIGHT
                step = -delta  # invert so up becomes left

                hbar = self.view.horizontalScrollBar()
                hbar.setValue(hbar.value() + step)
                return True

            return False  # allow normal vertical scrolling

        # ---------------------------
        # LEFT BUTTON PRESS
        # ---------------------------
        if event.type() == QEvent.Type.MouseButtonPress:
            mouse: QMouseEvent = event
            # ---------------------------
            # MIDDLE BUTTON = TEMP PAN/GRAB
            # ---------------------------
            if mouse.button() == Qt.MiddleButton:
                self.is_mid_panning = True
                self.mid_pan_start_pos = mouse.position().toPoint()
                self.mid_pan_start_h = self.view.horizontalScrollBar().value()
                self.mid_pan_start_v = self.view.verticalScrollBar().value()

                self.view.viewport().setCursor(Qt.ClosedHandCursor)
                return True

            # Right click cancels active tools
            if mouse.button() == Qt.RightButton:
                self.exit_tool_modes("Tools cancelled. Default mode active.")
                return True

            if mouse.button() == Qt.LeftButton:
                pos_scene = self.view.mapToScene(mouse.position().toPoint())
                # Exclusion zone: start drag
                if self.mode_exclusion_zone:
                    self.exclusion_drag_start_scene = pos_scene

                    if self.exclusion_drag_rect_item:
                        self.scene.removeItem(self.exclusion_drag_rect_item)
                        self.exclusion_drag_rect_item = None

                    self.exclusion_drag_rect_item = self.scene.addRect(
                        QRectF(pos_scene, pos_scene),
                        QPen(QColor(255, 0, 0), 2),
                        QBrush(QColor(255, 0, 0, 40))
                    )
                    self.exclusion_drag_rect_item.setZValue(50)
                    return True
                # Text-select: start drag
                if self.mode_text_select:
                    self.text_select_drag_start_scene = pos_scene

                    if self.text_select_rect_item:
                        self.scene.removeItem(self.text_select_rect_item)
                        self.text_select_rect_item = None

                    self.text_select_rect_item = self.scene.addRect(
                        QRectF(pos_scene, pos_scene),
                        QPen(QColor(0, 120, 255), 2),
                        QBrush(QColor(0, 120, 255, 40))
                    )
                    self.text_select_rect_item.setZValue(60)
                    return True

                # SHIFT + click dot = edit dialog (no dragging)
                if mouse.modifiers() & Qt.ShiftModifier:
                    dot_id = self.find_dot_item_near_scene(pos_scene, max_dist=18)
                    if dot_id is not None:
                        self.open_dot_edit_dialog(dot_id)
                        return True

                # If a tool mode is active, use your existing click logic
                if self.mode_pick_sample or self.mode_add_dot or self.mode_remove_dot:
                    self.on_scene_click(pos_scene)
                    return True

                # Default mode: click a dot and drag it
                dot_id = self.find_dot_item_near_scene(pos_scene, max_dist=18)
                if dot_id is not None:
                    self.dragging_dot_id = dot_id
                    self.is_dragging_dot = True

                    # offset so dot doesn't "snap" its center to cursor instantly
                    ellipse = self.dot_ellipse_items.get(dot_id)
                    if ellipse:
                        rect = ellipse.rect()
                        dot_center = QPointF(
                            rect.x() + rect.width() / 2.0,
                            rect.y() + rect.height() / 2.0
                        )
                        self.drag_dot_offset_scene = dot_center - pos_scene

                    self.statusBar().showMessage("Dragging dot… release to place.")
                    return True

                # Otherwise: allow normal pan/selection behavior
                return False

        # ---------------------------
        # MOUSE MOVE = drag dot
        # ---------------------------
        if event.type() == QEvent.Type.MouseMove:
            mouse: QMouseEvent = event
            # ---------------------------
            # MIDDLE BUTTON PAN DRAG
            # ---------------------------
            if self.is_mid_panning and (mouse.buttons() & Qt.MiddleButton):
                cur = mouse.position().toPoint()
                dx = cur.x() - self.mid_pan_start_pos.x()
                dy = cur.y() - self.mid_pan_start_pos.y()

                self.view.horizontalScrollBar().setValue(self.mid_pan_start_h - dx)
                self.view.verticalScrollBar().setValue(self.mid_pan_start_v - dy)
                return True
            pos_scene = self.view.mapToScene(mouse.position().toPoint())
            # Exclusion zone dragging: update rectangle
            if self.mode_exclusion_zone and self.exclusion_drag_start_scene and self.exclusion_drag_rect_item:
                r = QRectF(self.exclusion_drag_start_scene, pos_scene).normalized()
                self.exclusion_drag_rect_item.setRect(r)
                return True
            # Text-select dragging: update rectangle
            if self.mode_text_select and self.text_select_drag_start_scene and self.text_select_rect_item:
                r = QRectF(self.text_select_drag_start_scene, pos_scene).normalized()
                self.text_select_rect_item.setRect(r)
                return True

            self.last_hover_scene_pos = pos_scene
            # Feature 1: show dot-size previews while hovering in modes that need it
            if self.mode_pick_sample or self.mode_add_dot:
                self.set_dot_size_preview_visible(True)
                self.update_dot_size_previews()
            else:
                # don't leave overlays hanging in default mode
                if self._dot_size_preview_visible:
                    self.set_dot_size_preview_visible(False)


            # Hover magnifier for sampling (throttled to ~60fps)
            if self.mode_pick_sample:
                self._mag_pending_scene_pos = pos_scene
                if not self._mag_timer.isActive():
                    self._mag_timer.start()
                return False  # still allow movement
            if self.is_dragging_dot and self.dragging_dot_id is not None:
                pos_scene = self.view.mapToScene(mouse.position().toPoint())
                new_scene_pos = pos_scene + self.drag_dot_offset_scene
                self.move_dot_to_scene_pos(self.dragging_dot_id, new_scene_pos)
                return True
            return False

        # ---------------------------
        # LEFT RELEASE = stop dragging + persist label offsets
        # ---------------------------
        if event.type() == QEvent.Type.MouseButtonRelease:
            mouse: QMouseEvent = event
            # ---------------------------
            # MIDDLE BUTTON RELEASE = STOP PAN
            # ---------------------------
            if mouse.button() == Qt.MiddleButton and self.is_mid_panning:
                self.is_mid_panning = False

                # Restore cursor based on current mode
                if self.mode_pick_sample or self.mode_add_dot or self.mode_remove_dot or self.mode_exclusion_zone:
                    self.view.viewport().setCursor(Qt.CrossCursor)
                else:
                    self.view.viewport().unsetCursor()

                return True

            # Exclusion zone finish: store PDF rect
            if self.mode_exclusion_zone and self.exclusion_drag_start_scene and self.exclusion_drag_rect_item:
                final_rect_scene = self.exclusion_drag_rect_item.rect()

                p1 = final_rect_scene.topLeft()
                p2 = final_rect_scene.bottomRight()

                page1, x1, y1 = self.scene_to_page_px(p1)
                page2, x2, y2 = self.scene_to_page_px(p2)

                if page1 is not None and page1 == page2:
                    x1_pdf, y1_pdf = self.px_to_pdf(x1, y1)
                    x2_pdf, y2_pdf = self.px_to_pdf(x2, y2)

                    rect_pdf = QRectF(
                        QPointF(x1_pdf, y1_pdf),
                        QPointF(x2_pdf, y2_pdf)
                    ).normalized()

                    # 1) Save the exclusion zone (affects future detection)
                    self.exclusion_zones_by_page.setdefault(page1, []).append(rect_pdf)

                    # 2) NEW: remove any ALREADY-LABELED dots inside this zone
                    removed = self.remove_dots_inside_exclusion_rect(page1, rect_pdf)

                    # 3) NEW: renumber so numbers collapse correctly after removal
                    self.renumber_all_labels()

                    self.statusBar().showMessage(
                        f"Added exclusion zone on page {page1 + 1}. Removed {removed} dot(s) and renumbered."
                    )

                self.exclusion_drag_start_scene = None
                self.exclusion_drag_rect_item = None
                self.rebuild_overlay_items()
                self.renumber_all_labels()
                self.refresh_pages_list_ui()
                self.refresh_totals_ui()

                # EXIT exclusion mode after one rectangle
                self.exit_tool_modes("Exclusion zone added. Default mode active.")
                return True
            # Text-select finish: extract words and callback
            if self.mode_text_select and self.text_select_drag_start_scene and self.text_select_rect_item:
                final_rect_scene = self.text_select_rect_item.rect()

                p1 = final_rect_scene.topLeft()
                p2 = final_rect_scene.bottomRight()

                page1, x1, y1 = self.scene_to_page_px(p1)
                page2, x2, y2 = self.scene_to_page_px(p2)

                text = ""
                if page1 is not None and page1 == page2:
                    rect_px = QRectF(QPointF(x1, y1), QPointF(x2, y2)).normalized()
                    text = self.extract_pdf_text_in_rect(page1, rect_px)

                # cleanup
                try:
                    self.scene.removeItem(self.text_select_rect_item)
                except Exception:
                    pass
                self.text_select_rect_item = None
                self.text_select_drag_start_scene = None

                cb = self.text_select_callback
                self.exit_tool_modes("Text captured. Default mode active.")
                if cb:
                    cb(text)

                return True


            if mouse.button() == Qt.LeftButton:
                if self.is_dragging_dot:
                    self.is_dragging_dot = False
                    self.dragging_dot_id = None
                    self.statusBar().showMessage("Dot placed.")
                    return True

                # user may have dragged labels (they're movable) — update offsets
                self.update_label_offsets_from_scene()
                return False

        return super().eventFilter(source, event)

    def on_scene_click(self, pos: QPointF):
        if not self.doc:
            return

        page_index, x_px, y_px = self.scene_to_page_px(pos)
        if page_index is None:
            return

        if self.mode_pick_sample:
            self.pick_sample_at(page_index, x_px, y_px)
            return

        if self.mode_add_dot:
            self.add_dot_at(page_index, x_px, y_px)
            return

        if self.mode_remove_dot:
            dot_id = self.find_dot_item_near_scene(pos, max_dist=22)
            if dot_id is not None:
                self.open_dot_edit_dialog(dot_id)
            return

        # Default behavior: click a dot -> edit dialog
        dot_id = self.find_dot_item_near_scene(pos, max_dist=18)
        if dot_id is not None:
            self.open_dot_edit_dialog(dot_id)

    # ---------------- Coordinate Mapping ----------------

    def scene_to_page_px(self, pos_scene: QPointF) -> Tuple[Optional[int], float, float]:
        """
        Convert a scene coordinate into (doc_page_index, x_px, y_px) relative to that page's top-left.
        """
        x = float(pos_scene.x())
        y = float(pos_scene.y())

        for vis_i, y0 in enumerate(self.page_y_offsets):
            w, h = self.page_sizes_px[vis_i]
            if y0 <= y <= y0 + h and 0 <= x <= w:
                doc_page_index = self.visible_pages[vis_i]
                return doc_page_index, x, (y - y0)

        return None, 0.0, 0.0

    def page_px_to_scene(self, page_index: int, x_px: float, y_px: float) -> QPointF:
        """
        page_index is a DOC page index.
        """
        if page_index not in self.doc_to_visible_index:
            return QPointF(x_px, y_px)  # fallback, shouldn't occur if page is visible

        vis_i = self.doc_to_visible_index[page_index]
        y0 = self.page_y_offsets[vis_i]
        return QPointF(x_px, y0 + y_px)

    def px_to_pdf(self, x_px: float, y_px: float) -> Tuple[float, float]:
        return x_px / self.render_zoom, y_px / self.render_zoom

    def pdf_to_px(self, x_pdf: float, y_pdf: float) -> Tuple[float, float]:
        return x_pdf * self.render_zoom, y_pdf * self.render_zoom
    def update_magnifier(self, pos_scene: QPointF):
        """
        Shows a magnified view of the exact pixel under the cursor and displays BGR/HSV.
        Only used in Pick Sample mode.
        """
        if not self.doc:
            return

        page_index, x_px, y_px = self.scene_to_page_px(pos_scene)
        if page_index is None:
            self.magnifier.setVisible(False)
            self.magnifier_info.setVisible(False)
            return

        bgr = self.render_page_to_bgr(page_index)
        ix, iy = int(x_px), int(y_px)
        h, w = bgr.shape[:2]

        if ix < 0 or iy < 0 or ix >= w or iy >= h:
            return

        # sample 21x21 patch around the pixel
        half = 10
        x1, y1 = max(0, ix - half), max(0, iy - half)
        x2, y2 = min(w, ix + half + 1), min(h, iy + half + 1)
        patch = bgr[y1:y2, x1:x2].copy()

        # enlarge patch with nearest-neighbor scaling
        zoom = 6
        patch_big = cv2.resize(
            patch,
            (patch.shape[1] * zoom, patch.shape[0] * zoom),
            interpolation=cv2.INTER_NEAREST
        )

        # draw crosshair at center pixel
        cx = (ix - x1) * zoom
        cy = (iy - y1) * zoom
        cv2.drawMarker(
            patch_big,
            (cx, cy),
            (0, 0, 255),
            markerType=cv2.MARKER_CROSS,
            markerSize=18,
            thickness=2
        )

        # Convert to QPixmap
        rgb = cv2.cvtColor(patch_big, cv2.COLOR_BGR2RGB)
        qimg = QImage(rgb.data, rgb.shape[1], rgb.shape[0], rgb.strides[0], QImage.Format_RGB888)
        self.magnifier.setPixmap(
            QPixmap.fromImage(qimg.copy()).scaled(
                self.magnifier.size(),
                Qt.IgnoreAspectRatio,
                Qt.FastTransformation
            )
        )

        # Color info at exact pixel (no per-frame cvtColor call)
        b, g, r = bgr[iy, ix]
        hsv_img = self.render_page_to_hsv(page_index)
        hh, ss, vv = hsv_img[iy, ix]
        self.magnifier_info.setText(
            f"PX ({ix},{iy})  BGR=({b},{g},{r})  HSV=({int(hh)},{int(ss)},{int(vv)})"
        )

        # Position magnifier near the cursor in main-window coords
        cursor_pos = QCursor.pos()  # GLOBAL screen coords
        self.magnifier.move(cursor_pos.x() + 20, cursor_pos.y() + 20)

        # Always show + keep on top
        self.magnifier.show()
        self.magnifier.raise_()


    def _nearest_point_on_rect(self, rect: QRectF, p: QPointF) -> QPointF:
        """
        Clamp p to the nearest point on/in rect (axis-aligned).
        """
        x = min(max(p.x(), rect.left()), rect.right())
        y = min(max(p.y(), rect.top()), rect.bottom())
        return QPointF(x, y)

    def _point_on_circle_towards(self, center: QPointF, radius: float, target: QPointF) -> QPointF:
        """
        Point on circle perimeter starting at center going towards target.
        """
        dx = target.x() - center.x()
        dy = target.y() - center.y()
        dist = math.hypot(dx, dy)
        if dist <= 1e-6:
            return QPointF(center.x() + radius, center.y())
        ux = dx / dist
        uy = dy / dist
        return QPointF(center.x() + ux * radius, center.y() + uy * radius)

    def update_connector_for_dot(self, dot_id: int):
        """
        Ensure a connector line exists and is positioned from dot perimeter
        to the nearest point on the label box.
        """
        ellipse = self.dot_ellipse_items.get(dot_id)
        label = self.dot_label_items.get(dot_id)
        if ellipse is None or label is None:
            # If either is missing, remove the connector if it exists
            old = self.dot_connector_items.pop(dot_id, None)
            if old is not None:
                try:
                    self.scene.removeItem(old)
                except Exception:
                    pass
            return

        # Dot center + radius in scene coords
        er = ellipse.rect()
        dot_center = QPointF(er.x() + er.width() / 2.0, er.y() + er.height() / 2.0)
        dot_radius = er.width() / 2.0

        # Label box rect in *scene* coords
        label_scene_rect = label.mapToScene(label.boundingRect()).boundingRect()

        # Nearest point on label rect to dot center
        label_attach = self._nearest_point_on_rect(label_scene_rect, dot_center)

        # Point on dot perimeter towards label_attach
        dot_attach = self._point_on_circle_towards(dot_center, dot_radius, label_attach)

        # Color = model sample color
        page_i, dot_i = self.dot_index[dot_id]
        dot_obj = self.dots_by_page[page_i][dot_i]
        c = self.model_color(dot_obj.model)

        pen = QPen(QColor(c), max(1, int(round(2 * self.label_scale_factor()))))
        pen.setCapStyle(Qt.RoundCap)

        line_item = self.dot_connector_items.get(dot_id)
        if line_item is None:
            line_item = QGraphicsLineItem()
            line_item.setZValue(15)  # above dot (10), below label (20)
            self.scene.addItem(line_item)
            self.dot_connector_items[dot_id] = line_item

        line_item.setPen(pen)
        line_item.setLine(dot_attach.x(), dot_attach.y(), label_attach.x(), label_attach.y())

    def rebuild_all_connectors(self):
        """
        Rebuild/update connectors for all labels.
        """
        # Remove connectors for dot_ids that no longer exist
        valid_ids = set(self.dot_label_items.keys())
        for dot_id in list(self.dot_connector_items.keys()):
            if dot_id not in valid_ids:
                old = self.dot_connector_items.pop(dot_id, None)
                if old is not None:
                    try:
                        self.scene.removeItem(old)
                    except Exception:
                        pass

        # Update/create connectors for all label items
        for dot_id in self.dot_label_items.keys():
            self.update_connector_for_dot(dot_id)


    def model_color(self, model: str) -> QColor:
        """
        Returns a QColor for the model based on its sampled HSV.
        Used for row highlighting + halo overlays.
        """
        if model in self.model_colors:
            return self.model_colors[model]

        sample = self.samples.get(model)
        if not sample:
            c = QColor(255, 0, 255)  # fallback magenta
            self.model_colors[model] = c
            return c

        h, s, v = sample.hsv
        hsv_px = np.uint8([[[h, s, v]]])  # OpenCV HSV
        rgb = cv2.cvtColor(hsv_px, cv2.COLOR_HSV2RGB)[0][0]
        c = QColor(int(rgb[0]), int(rgb[1]), int(rgb[2]))
        self.model_colors[model] = c
        return c

    def clear_model_highlight(self):
        self.active_highlight_model = None
        self.clear_halos()
        self.refresh_totals_ui()

    def clear_halos(self):
        # Stop animation if nothing should be breathing
        if self.halo_anim_timer.isActive():
            self.halo_anim_timer.stop()

        for it in self.halo_items:
            try:
                self.scene.removeItem(it)
            except Exception:
                pass
        self.halo_items = []

    def apply_model_halos(self, model: str):
        """
        Adds a 50% transparent halo around all dots of this model across all visible pages.
        """
        self.clear_halos()

        if not model or model not in self.get_model_list():
            return

        color = self.model_color(model)
        halo_brush = QColor(color)
        halo_brush.setAlpha(128)  # 50% transparency

        for page_index in self.visible_pages:
            dots = self.dots_by_page.get(page_index, [])
            for d in dots:
                if d.excluded or d.model != model:
                    continue

                cx_px, cy_px = self.pdf_to_px(d.cx_pdf, d.cy_pdf)
                scene_pos = self.page_px_to_scene(page_index, cx_px, cy_px)

                base_r_px = max(6.0, d.radius_pdf * self.render_zoom)

                # NEW: DOUBLE the halo size (previously 1.65x)
                halo_r = base_r_px * 3.30

                halo = QGraphicsEllipseItem(
                    scene_pos.x() - halo_r,
                    scene_pos.y() - halo_r,
                    halo_r * 2,
                    halo_r * 2
                )

                halo.setBrush(QBrush(halo_brush))
                halo.setPen(QPen(QColor(0, 0, 0, 0), 0))
                halo.setZValue(8)  # behind dot outline (dot ellipse uses Z=10)
                self.scene.addItem(halo)
                self.halo_items.append(halo)
        # NEW: start breathing animation
        self.halo_anim_clock.restart()
        if not self.halo_anim_timer.isActive():
            self.halo_anim_timer.start()

    def _tick_halo_breathe(self):
        """
        Opacity oscillates 100% -> 0% -> 100% every 1 second.
        Uses a cosine wave for smooth breathing.
        """
        if not self.halo_items and not self.focus_halo_items:
            return

        t = self.halo_anim_clock.elapsed()  # ms
        phase = (t % self.halo_breathe_period_ms) / float(self.halo_breathe_period_ms)

        # cosine wave: 1 -> -1 -> 1  => mapped to 255 -> 0 -> 255
        a = int(((math.cos(phase * 2.0 * math.pi) + 1.0) * 0.5) * 255)

        for halo in (self.halo_items + self.focus_halo_items):
            brush = halo.brush()
            c = brush.color()
            c.setAlpha(a)
            brush.setColor(c)
            halo.setBrush(brush)

    def clear_focus_halo(self):
        for it in self.focus_halo_items:
            try:
                self.scene.removeItem(it)
            except Exception:
                pass
        self.focus_halo_items = []

    def apply_focus_halo_to_dot(self, dot: Dot):
        """
        Breathing halo around ONE dot (current naming target).
        """
        self.clear_focus_halo()

        if not dot or dot.excluded:
            return

        # center in scene
        cx_px, cy_px = self.pdf_to_px(dot.cx_pdf, dot.cy_pdf)
        scene_pos = self.page_px_to_scene(dot.page_index, cx_px, cy_px)

        base_r_px = max(6.0, dot.radius_pdf * self.render_zoom)
        halo_r = base_r_px * 3.3

        c = self.model_color(dot.model)
        halo_brush = QColor(c)
        halo_brush.setAlpha(180)

        halo = QGraphicsEllipseItem(
            scene_pos.x() - halo_r,
            scene_pos.y() - halo_r,
            halo_r * 2,
            halo_r * 2
        )
        halo.setBrush(QBrush(halo_brush))
        halo.setPen(QPen(QColor(0, 0, 0, 0), 0))
        halo.setZValue(9)
        self.scene.addItem(halo)
        self.focus_halo_items.append(halo)

        # start the existing breathe timer if needed
        self.halo_anim_clock.restart()
        if not self.halo_anim_timer.isActive():
            self.halo_anim_timer.start()

    def get_dot_by_uid(self, uid: str) -> Optional[Dot]:
        for pi, dots in self.dots_by_page.items():
            for d in dots:
                if d.uid == uid:
                    return d
        return None

    def focus_dot(self, uid: str, pad_px: float = 350.0):
        """
        Pan/zoom to the dot so the user can read nearby context, and make ONLY that dot breathe.
        Used by the naming dialog (Name Devices... / Next) and anywhere else you want single-dot focus.
        """
        self.focus_dot_zoomed(uid, pad_px=pad_px)

    # ---------------- Samples / Model Order ----------------

    def pick_sample_at(self, page_index: int, x_px: float, y_px: float):
        bgr = self.render_page_to_bgr(page_index)
        hsv = cv2.cvtColor(bgr, cv2.COLOR_BGR2HSV)

        # dot radius guess for sampling at current zoom
        ring = sample_dot_hsv(hsv, int(x_px), int(y_px))
        if ring is None:
            QMessageBox.information(self, "Sample", "Could not sample a dot color here. Click more within the dot ring.")
            return

        # Hide magnifier so dialog is never "behind" tooltip overlays
        was_showing_mag = self.magnifier.isVisible() or self.magnifier_info.isVisible()
        self.magnifier.hide()
        self.magnifier_info.hide()

        dlg = QInputDialog(self)
        dlg.setWindowTitle("Model Name")
        dlg.setLabelText("Enter model name (e.g. C11, C1, C3):")
        dlg.setInputMode(QInputDialog.TextInput)
        dlg.setTextValue("")
        dlg.setWindowModality(Qt.ApplicationModal)
        dlg.setOkButtonText("OK (`)")
        dlg.setCancelButtonText("Cancel")

        # ---------------------------------------------------
        # OK button text + ` shortcut + preset buttons (2x5)
        # ---------------------------------------------------

        # Ensure button labels are correct (you already set these above, but keep them here for safety)
        dlg.setOkButtonText("OK (`)")
        dlg.setCancelButtonText("Cancel")

        # QShortcut for ` (works even when typing)
        sc_backtick = QShortcut(QKeySequence("`"), dlg)
        sc_backtick.setContext(Qt.WidgetWithChildrenShortcut)
        sc_backtick.activated.connect(dlg.accept)

        # Fallback: event-filter catches layouts where QKeySequence("`") may not fire
        class _BacktickAcceptFilter(QObject):
            def eventFilter(self, obj, event):
                if event.type() == QEvent.Type.KeyPress:
                    if event.key() == Qt.Key_QuoteLeft:  # physical ` / ~ key
                        dlg.accept()
                        return True
                return False

        _bt_filter = _BacktickAcceptFilter(dlg)
        dlg.installEventFilter(_bt_filter)

        le = dlg.findChild(QLineEdit)
        if le:
            le.installEventFilter(_bt_filter)

        # ---------- Preset buttons (2 rows x 5) ----------
        presets = ["C1", "C2", "C3", "C4", "C5",
                   "C10", "C11", "C14", "M-1", "M-2"]

        preset_widget = QWidget(dlg)
        grid = QGridLayout(preset_widget)
        grid.setContentsMargins(0, 6, 0, 6)  # a little breathing room
        grid.setHorizontalSpacing(6)
        grid.setVerticalSpacing(6)

        def _choose_preset(txt: str):
            dlg.setTextValue(txt)
            dlg.accept()  # same effect as pressing Enter / OK

        for i, txt in enumerate(presets):
            b = QPushButton(txt, preset_widget)
            b.setMinimumHeight(28)
            grid.addWidget(b, i // 5, i % 5)
            b.clicked.connect(lambda checked=False, t=txt: _choose_preset(t))

        # Insert the preset button grid ABOVE the OK/Cancel button row
        bb = dlg.findChild(QDialogButtonBox)
        lay = dlg.layout()
        if lay and bb:
            lay.insertWidget(lay.indexOf(bb), preset_widget)
        elif lay:
            lay.addWidget(preset_widget)
        # ---------------------------------------------------

        # Force it visually above any tooltips/magnifier windows
        dlg.show()
        dlg.raise_()
        dlg.activateWindow()

        ok = (dlg.exec() == QDialog.Accepted)
        model = dlg.textValue()

        if not ok or not model.strip():
            # Restore magnifier if still in pick-sample mode
            if was_showing_mag and self.mode_pick_sample:
                self.magnifier.show()
            return

        model = model.strip().upper()

        # Restore magnifier if still in pick-sample mode
        if was_showing_mag and self.mode_pick_sample:
            self.magnifier.show()

        self.samples[model] = ModelSample(model=model, hsv=ring)

        # REPLACE: auto-place newly sampled models before higher-numbered siblings (e.g. C4 before C11)
        self._auto_insert_model_order(model)

        self.refresh_samples_ui()
        self.refresh_model_order_ui()

        self.statusBar().showMessage(f"Sample saved: {model} HSV={ring}")

        # Immediately run detection + label updates for ALL pages
        for pi in self.visible_pages:
            self.detect_page(pi)
        self.renumber_all_labels()
        self.rebuild_overlay_items()
        self.refresh_pages_list_ui()
        self.refresh_totals_ui()

    def label_scale_factor(self) -> float:
        """
        Slider pct is -100..+200 where 0 => 1.0x.
        Clamp to >0 to avoid invalid font sizes.
        """
        return max(0.05, 1.0 + (float(self.label_scale_pct) / 100.0))

    def on_label_scale_changed(self, value: int):
        """
        Update label scale and immediately apply to all existing label items.
        Also updates the status text.
        """
        self.label_scale_pct = int(value)
        pct = 100 + self.label_scale_pct
        self.lbl_label_scale.setText(f"Label Size: {pct}%")

        # Apply to existing on-screen labels
        scale = self.label_scale_factor()
        for label_item in self.dot_label_items.values():
            if hasattr(label_item, "set_scale"):
                label_item.set_scale(scale)

        # Force scene redraw (helps if some viewports don't repaint immediately)
        self.scene.update()
        self.view.viewport().update()

    def refresh_samples_ui(self):
        self.sample_table.setRowCount(0)
        for model in sorted(self.samples.keys()):
            row = self.sample_table.rowCount()
            self.sample_table.insertRow(row)
            self.sample_table.setItem(row, 0, QTableWidgetItem(model))
            self.sample_table.setItem(row, 1, QTableWidgetItem(str(self.samples[model].hsv)))

    def refresh_model_order_ui(self):
        # Model order is now managed by drag/drop in the ALL totals table.
        # Keep this method so existing calls don't break.
        self.refresh_totals_ui()

    # ---------------------------
    # Floors / Totals UI
    # ---------------------------

    def get_floor_title(self, doc_page_index: int) -> str:
        """
        Returns the display title for a doc page index.
        Default = 'Floor N'
        """
        if doc_page_index in self.page_titles and self.page_titles[doc_page_index].strip():
            return self.page_titles[doc_page_index].strip()
        return f"Floor {doc_page_index + 1}"

    def refresh_pages_list_ui(self):
        """
        Populate the right-side page list with visible pages.
        """
        self.page_list.clear()

        if not self.doc:
            return

        for doc_pi in self.visible_pages:
            title = self.get_floor_title(doc_pi)
            item = QListWidgetItem(title)
            item.setData(Qt.UserRole, doc_pi)
            self.page_list.addItem(item)

    def refresh_totals_ui(self):
        """
        Builds totals tabs.

        NEW behavior:
        - Always creates: ALL tab + Floor tabs for every visible page (even if 0 dots)
        - ALL tab totals across all pages
        - Clicking a model row toggles highlight + halos
        """
        # Remember current tab title so we can restore it after rebuild
        prev_tab_text = None
        if self.totals_tabs.count() > 0 and self.totals_tabs.currentIndex() >= 0:
            prev_tab_text = self.totals_tabs.tabText(self.totals_tabs.currentIndex())

        self.totals_tabs.clear()

        if not self.doc:
            return

        models = self.get_model_list()

        def build_counts_for_page(doc_pi: Optional[int]) -> Dict[str, int]:
            counts = {m: 0 for m in models}

            if doc_pi is None:
                # ALL pages
                for pi in self.visible_pages:
                    for d in self.dots_by_page.get(pi, []):
                        if d.excluded:
                            continue
                        counts[d.model] = counts.get(d.model, 0) + 1
            else:
                # One page
                for d in self.dots_by_page.get(doc_pi, []):
                    if d.excluded:
                        continue
                    counts[d.model] = counts.get(d.model, 0) + 1

            return counts

        def make_table(doc_pi: Optional[int]) -> QWidget:
            counts = build_counts_for_page(doc_pi)

            # ---------------------------
            # Main scrolling table
            # ---------------------------
            if doc_pi is None:
                # ALL tab: allow row reordering (above/below only)
                table = ModelReorderTable(0, 2,
                                          on_reorder=lambda t_ref=None: self.sync_model_order_from_totals_table(table))
            else:
                table = QTableWidget(0, 2)

            table.setHorizontalHeaderLabels(["Model", "Count"])
            table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
            table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
            table.verticalHeader().setVisible(False)
            table.setEditTriggers(QAbstractItemView.NoEditTriggers)
            table.setSelectionBehavior(QAbstractItemView.SelectRows)
            table.setSelectionMode(QAbstractItemView.SingleSelection)

            for m in models:
                row = table.rowCount()
                table.insertRow(row)

                it_model = QTableWidgetItem(m)
                it_count = QTableWidgetItem(str(counts.get(m, 0)))

                # Visual emphasis when active highlight matches this model
                if self.active_highlight_model and self.active_highlight_model == m:
                    c = self.model_color(m)
                    it_model.setBackground(QBrush(QColor(c.red(), c.green(), c.blue(), 70)))
                    it_model.setFont(QFont("Arial", 10, QFont.Bold))
                    it_count.setBackground(QBrush(QColor(c.red(), c.green(), c.blue(), 40)))
                    it_count.setFont(QFont("Arial", 10, QFont.Bold))

                table.setItem(row, 0, it_model)
                table.setItem(row, 1, it_count)
            # NEW: allow drag/drop reorder ONLY on the ALL tab
            if doc_pi is None:
                self._enable_model_reorder_on_table(table)

            table.cellClicked.connect(lambda r, col, p=doc_pi, t=table: self.on_totals_model_clicked(p, t, r))

            # ---------------------------
            # Frozen TOTAL footer bar
            # ---------------------------
            total = sum(int(counts.get(m, 0)) for m in models)

            footer = QTableWidget(1, 2)
            footer.setHorizontalHeaderLabels(["Model", "Count"])
            footer.verticalHeader().setVisible(False)
            footer.horizontalHeader().setVisible(False)
            footer.setEditTriggers(QAbstractItemView.NoEditTriggers)
            footer.setSelectionMode(QAbstractItemView.NoSelection)
            footer.setFocusPolicy(Qt.NoFocus)

            it_total = QTableWidgetItem("TOTAL")
            it_total_count = QTableWidgetItem(str(total))

            bold = QFont("Arial", 10)
            bold.setBold(True)
            it_total.setFont(bold)
            it_total_count.setFont(bold)

            footer.setItem(0, 0, it_total)
            footer.setItem(0, 1, it_total_count)

            # Make footer visually match the main table (dark mode / stylesheet / palette)
            footer.setAutoFillBackground(True)
            footer.setPalette(table.palette())
            footer.viewport().setPalette(table.viewport().palette())
            footer.setStyleSheet(table.styleSheet())

            # Column sizing:
            # - Column 0 wide enough to show "TOTAL"
            # - Column 1 stretches like the main table
            footer.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
            footer.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
            footer.resizeColumnToContents(0)

            # Match header widths so columns line up perfectly
            footer.setColumnWidth(0, table.columnWidth(0))
            footer.setColumnWidth(1, table.columnWidth(1))

            # Keep footer compact (one-row bar)
            footer.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
            footer.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
            footer.setFixedHeight(footer.rowHeight(0) + 6)

            # If the user resizes columns (or Qt changes them), keep footer aligned
            table.horizontalHeader().sectionResized.connect(
                lambda idx, old, new, f=footer: f.setColumnWidth(idx, new)
            )

            return TotalsTabWidget(table, footer)

        # Tab 0: ALL
        all_tab = make_table(None)
        self.totals_tabs.addTab(all_tab, "ALL")

        # Tab 1: NAMES (device table)
        names_table = self.make_names_table_widget()
        self.totals_tabs.addTab(names_table, "NAMES")

        # Floor tabs: always exist
        for doc_pi in self.visible_pages:
            tab_title = self.get_floor_title(doc_pi)
            t_tab = make_table(doc_pi)
            self.totals_tabs.addTab(t_tab, tab_title)

        # Restore tab selection or force NAMES during naming mode
        if self.force_names_tab_selected:
            # NAMES is always tab index 1 (ALL is 0)
            if self.totals_tabs.count() > 1:
                self.totals_tabs.setCurrentIndex(1)
        elif prev_tab_text:
            for i in range(self.totals_tabs.count()):
                if self.totals_tabs.tabText(i) == prev_tab_text:
                    self.totals_tabs.setCurrentIndex(i)
                    break

        # Re-apply highlight ONLY when it makes sense:
        # - naming is active (force_names_tab_selected), OR
        # - user is already on the NAMES tab
        current_tab_text = None
        if self.totals_tabs.count() > 0 and self.totals_tabs.currentIndex() >= 0:
            current_tab_text = self.totals_tabs.tabText(self.totals_tabs.currentIndex())

        if self.names_highlight_uid and (self.force_names_tab_selected or current_tab_text == "NAMES"):
            self.highlight_names_uid(self.names_highlight_uid)

    def rename_floor_tab(self, tab_index: int):
        """
        Double-click a totals tab to rename it.
        """
        if tab_index < 0:
            return

        # We need to know which doc page this tab corresponds to:
        # We'll rebuild the same list used in refresh_totals_ui
        tab_doc_pages = []
        for doc_pi in self.visible_pages:
            dots = self.dots_by_page.get(doc_pi, [])
            dots = [d for d in dots if not d.excluded]
            if dots:
                tab_doc_pages.append(doc_pi)

        if tab_index >= len(tab_doc_pages):
            return

        doc_pi = tab_doc_pages[tab_index]
        current = self.get_floor_title(doc_pi)

        new_name, ok = QInputDialog.getText(self, "Rename Floor", "Enter new Floor name:", text=current)
        if not ok or not new_name.strip():
            return

        self.page_titles[doc_pi] = new_name.strip()
        self.refresh_pages_list_ui()
        self.refresh_totals_ui()

    def on_totals_model_clicked(self, doc_pi: Optional[int], table: QTableWidget, row: int):
        """
        Clicking a row toggles the model highlight.
        - Click same model again => turn off
        - Selecting another model => switch highlight
        """
        item = table.item(row, 0)
        if not item:
            return

        model = item.text().strip()

        # Toggle off if same model already active
        if self.active_highlight_model == model:
            self.clear_model_highlight()
            return

        self.active_highlight_model = model
        self.apply_model_halos(model)
        self.refresh_totals_ui()


    def rename_selected_floor(self):
        """
        Rename via the list + button.
        """
        item = self.page_list.currentItem()
        if not item:
            return

        doc_pi = item.data(Qt.UserRole)
        current = self.get_floor_title(doc_pi)

        new_name, ok = QInputDialog.getText(self, "Rename Floor", "Enter new Floor name:", text=current)
        if not ok or not new_name.strip():
            return

        self.page_titles[doc_pi] = new_name.strip()
        self.refresh_pages_list_ui()
        self.refresh_totals_ui()

    def delete_selected_floor(self):
        """
        Remove a page from the program (without touching the underlying PDF file).
        """
        if not self.doc:
            return

        item = self.page_list.currentItem()
        if not item:
            return

        doc_pi = item.data(Qt.UserRole)

        confirm = QMessageBox.question(
            self,
            "Delete Floor",
            f"Remove {self.get_floor_title(doc_pi)} from the program?\n\n"
            "This does NOT delete the PDF file.\n"
            "It just removes this page from labeling, totals, and exports.",
            QMessageBox.Yes | QMessageBox.No
        )
        if confirm != QMessageBox.Yes:
            return

        self.removed_pages.add(int(doc_pi))

        # Also drop any detected dots/exclusions for that page
        if doc_pi in self.dots_by_page:
            self.dots_by_page.pop(doc_pi, None)
        if doc_pi in self.exclusion_zones_by_page:
            self.exclusion_zones_by_page.pop(doc_pi, None)

        self.render_all_pages()
        self.renumber_all_labels()
        self.rebuild_overlay_items()
        self.refresh_pages_list_ui()
        self.refresh_totals_ui()

    # ---------------------------
    # Export Totals + Devices Tables
    # ---------------------------

    def export_totals_and_devices_tables(self):
        """
        Exports:
        1) The currently-selected totals tab table (Model/Count, headers included)
        2) A per-device table: Number | Camera Type | Name  (Name defaults to label, e.g. 1-C1)

        Output format options:
        A) Excel (.xlsx)  -> 2 sheets: Totals, Devices
        B) CSV (.csv)     -> 2 files: *_totals.csv and *_devices.csv
        C) PNG (.png)     -> 2 files: *_totals.png and *_devices.png
        D) Append to PDF  -> creates a NEW pdf with an extra appended page containing both tables
        """
        if not self.doc:
            return

        # Grab the currently selected totals tab widget
        current = self.totals_tabs.currentWidget()
        if current is None:
            QMessageBox.information(self, "Export Table", "No totals table is currently available to export.")
            return

        # NEW: totals tabs may be a container with .main_table
        if isinstance(current, QTableWidget):
            totals_table = current
        elif hasattr(current, "main_table") and isinstance(current.main_table, QTableWidget):  # type: ignore
            totals_table = current.main_table  # type: ignore
        else:
            QMessageBox.information(self, "Export Table", "No totals table is currently available to export.")
            return

        totals_headers, totals_rows = self._extract_qtable(totals_table)

        devices_headers, devices_rows = self._build_devices_rows()

        # Choose export format
        options = [
            "Excel (.xlsx)",
            "CSV (.csv)",
            "Image (.png)",
            "Append to end of PDF (new PDF file)"
        ]
        choice, ok = QInputDialog.getItem(
            self, "Export Table", "Choose export format:", options, 0, False
        )
        if not ok or not choice:
            return

        if choice.startswith("Excel"):
            path, _ = QFileDialog.getSaveFileName(self, "Save Excel", "", "Excel Workbook (*.xlsx)")
            if not path:
                return
            if not path.lower().endswith(".xlsx"):
                path += ".xlsx"
            self._export_xlsx(path, totals_headers, totals_rows, devices_headers, devices_rows)
            QMessageBox.information(self, "Export", f"Saved Excel:\n{path}")
            return

        if choice.startswith("CSV"):
            path, _ = QFileDialog.getSaveFileName(self, "Save CSV (base name)", "", "CSV (*.csv)")
            if not path:
                return
            # We'll create two CSVs using this as the base name
            base = path[:-4] if path.lower().endswith(".csv") else path
            totals_path = f"{base}_totals.csv"
            devices_path = f"{base}_devices.csv"
            self._export_csv(totals_path, totals_headers, totals_rows)
            self._export_csv(devices_path, devices_headers, devices_rows)
            QMessageBox.information(self, "Export", f"Saved CSVs:\n{totals_path}\n{devices_path}")
            return

        if choice.startswith("Image"):
            path, _ = QFileDialog.getSaveFileName(self, "Save PNG (base name)", "", "PNG (*.png)")
            if not path:
                return
            base = path[:-4] if path.lower().endswith(".png") else path
            totals_path = f"{base}_totals.png"
            devices_path = f"{base}_devices.png"
            self._export_png_from_qtable(totals_table, totals_path)

            devices_table = self._make_devices_table_widget(devices_headers, devices_rows)
            self._export_png_from_qtable(devices_table, devices_path)

            QMessageBox.information(self, "Export", f"Saved PNGs:\n{totals_path}\n{devices_path}")
            return

        # Append to PDF (new file)
        out_path, _ = QFileDialog.getSaveFileName(self, "Save PDF (with appended tables)", "", "PDF Files (*.pdf)")
        if not out_path:
            return
        if not out_path.lower().endswith(".pdf"):
            out_path += ".pdf"

        ok = self._export_append_pdf_with_tables(
            out_path,
            totals_headers, totals_rows,
            devices_headers, devices_rows
        )
        if ok:
            QMessageBox.information(self, "Export", f"Saved PDF:\n{out_path}")


    def _extract_qtable(self, table: QTableWidget) -> Tuple[List[str], List[List[str]]]:
        """
        Extract headers + all rows (as strings), including 0-count rows.
        """
        headers = []
        for c in range(table.columnCount()):
            h = table.horizontalHeaderItem(c)
            headers.append(h.text() if h else f"Col{c+1}")

        rows = []
        for r in range(table.rowCount()):
            row = []
            for c in range(table.columnCount()):
                it = table.item(r, c)
                row.append(it.text() if it else "")
            rows.append(row)

        return headers, rows


    def _safe_parse_label_number(self, label: str) -> int:
        """
        label like "12-C5" -> 12, else large number to push unknowns to end.
        """
        try:
            left = label.split("-", 1)[0].strip()
            return int(left)
        except Exception:
            return 10**9


    def _build_devices_rows(self) -> Tuple[List[str], List[List[str]]]:
        """
        Builds:
          Number | Camera Type | Name
        where Name defaults to the dot label (e.g. 1-C1).
        Sorted by Number.
        """
        headers = ["Number", "Camera Type", "Name"]
        rows = []

        for page_index in self.visible_pages:
            for d in self.dots_by_page.get(page_index, []):
                if d.excluded:
                    continue

                label = d.label or ""
                # Model from dot (fallback: parse from label)
                cam_type = d.model or (label.split("-", 1)[1] if "-" in label else "")

                num = self._safe_parse_label_number(label) if label else 10**9
                rows.append([str(num if num != 10**9 else ""), cam_type, label])

        # Sort by numeric number
        rows.sort(key=lambda r: int(r[0]) if str(r[0]).isdigit() else 10**9)
        return headers, rows


    def _export_csv(self, path: str, headers: List[str], rows: List[List[str]]):
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(headers)
            for r in rows:
                w.writerow(r)


    def _export_xlsx(
        self,
        path: str,
        totals_headers: List[str],
        totals_rows: List[List[str]],
        devices_headers: List[str],
        devices_rows: List[List[str]]
    ):
        wb = Workbook()

        # Sheet 1: Totals
        ws1 = wb.active
        ws1.title = "Totals"
        self._write_sheet(ws1, totals_headers, totals_rows)

        # Sheet 2: Devices
        ws2 = wb.create_sheet("Devices")
        self._write_sheet(ws2, devices_headers, devices_rows)

        wb.save(path)


    def _write_sheet(self, ws, headers: List[str], rows: List[List[str]]):
        bold = Font(bold=True)
        center = Alignment(horizontal="center")

        ws.append(headers)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = bold
            cell.alignment = center

        for r in rows:
            ws.append(r)

        # Simple column width autosize
        for col in range(1, len(headers) + 1):
            max_len = 0
            for row in range(1, ws.max_row + 1):
                v = ws.cell(row=row, column=col).value
                if v is None:
                    continue
                max_len = max(max_len, len(str(v)))
            ws.column_dimensions[chr(64 + col)].width = min(max_len + 3, 60)


    def _make_devices_table_widget(self, headers: List[str], rows: List[List[str]]) -> QTableWidget:
        """
        Creates a temporary QTableWidget for rendering/export to PNG.
        """
        t = QTableWidget(len(rows), len(headers))
        t.setHorizontalHeaderLabels(headers)
        t.verticalHeader().setVisible(False)
        t.setEditTriggers(QAbstractItemView.NoEditTriggers)

        for r, row in enumerate(rows):
            for c, val in enumerate(row):
                t.setItem(r, c, QTableWidgetItem(str(val)))

        t.resizeColumnsToContents()
        t.resizeRowsToContents()

        # Make sure the widget is sized to fit its contents before grabbing
        w = t.horizontalHeader().length() + t.frameWidth() * 2 + 20
        h = t.verticalHeader().length() + t.horizontalHeader().height() + t.frameWidth() * 2 + 20
        t.setFixedSize(max(300, w), max(200, h))
        return t


    def _export_png_from_qtable(self, table: QTableWidget, path: str):
        """
        Renders a QTableWidget to a PNG.
        """
        # Ensure it has a reasonable size
        table.resizeColumnsToContents()
        table.resizeRowsToContents()

        pix = table.grab()
        pix.save(path, "PNG")


    def _export_append_pdf_with_tables(
        self,
        out_path: str,
        totals_headers: List[str],
        totals_rows: List[List[str]],
        devices_headers: List[str],
        devices_rows: List[List[str]]
    ) -> bool:
        """
        Creates a NEW PDF = original PDF + appended page(s) containing the totals table and devices table.
        """
        if not self.pdf_path:
            QMessageBox.critical(self, "Export Error", "No source PDF path is available.")
            return False

        try:
            doc = fitz.open(self.pdf_path)

            # Choose a page size matching the first page if possible
            if doc.page_count > 0:
                rect = doc[0].rect
                page_w, page_h = rect.width, rect.height
            else:
                page_w, page_h = 612, 792  # fallback

            # Helper to add pages as needed
            def new_page():
                return doc.new_page(width=page_w, height=page_h)

            page = new_page()
            y = 40
            x = 40
            line_h = 14
            font = "courier"
            font_size = 10

            def draw_title(p, text, y_pos):
                p.insert_text((x, y_pos), text, fontsize=12, fontname="helv", color=(0, 0, 0))
                return y_pos + 22

            def draw_table(p, title, headers, rows, y_pos):
                nonlocal page
                y_pos = draw_title(p, title, y_pos)

                # column widths in characters (monospace) – tuned for your specific tables
                # Totals: Model | Count
                # Devices: Number | Camera Type | Name
                if len(headers) == 2:
                    col_chars = [14, 10]
                else:
                    col_chars = [10, 14, 60]

                def fmt_row(cols):
                    out = ""
                    for i, c in enumerate(cols):
                        s = str(c)
                        w = col_chars[i]
                        if len(s) > w - 1:
                            s = s[: w - 4] + "..."
                        out += s.ljust(w)
                    return out

                # Header
                header_line = fmt_row(headers)
                sep_line = "-" * min(len(header_line), 95)

                # Page break check
                if y_pos + line_h * 3 > page_h - 40:
                    page = new_page()
                    p = page
                    y_pos = 40

                p.insert_text((x, y_pos), header_line, fontsize=font_size, fontname=font, color=(0, 0, 0))
                y_pos += line_h
                p.insert_text((x, y_pos), sep_line, fontsize=font_size, fontname=font, color=(0, 0, 0))
                y_pos += line_h

                # Rows
                for r in rows:
                    if y_pos + line_h > page_h - 40:
                        page = new_page()
                        p = page
                        y_pos = 40
                        # repeat header on new page
                        p.insert_text((x, y_pos), header_line, fontsize=font_size, fontname=font, color=(0, 0, 0))
                        y_pos += line_h
                        p.insert_text((x, y_pos), sep_line, fontsize=font_size, fontname=font, color=(0, 0, 0))
                        y_pos += line_h

                    p.insert_text((x, y_pos), fmt_row(r), fontsize=font_size, fontname=font, color=(0, 0, 0))
                    y_pos += line_h

                return y_pos + 18, p

            # Draw Totals then Devices
            y, page = draw_table(page, "Totals (Current Tab)", totals_headers, totals_rows, y)
            y, page = draw_table(page, "Devices", devices_headers, devices_rows, y)

            doc.save(out_path)
            doc.close()
            return True

        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to append tables to PDF:\n{e}")
            return False

    def _enable_model_reorder_on_table(self, table: QTableWidget):
        """
        Turns the Model|Count table into a drag/drop reorder list (row moves).
        """
        table.setDragEnabled(True)
        table.setAcceptDrops(True)
        table.setDropIndicatorShown(True)
        table.setDragDropOverwriteMode(False)
        table.setDragDropMode(QAbstractItemView.InternalMove)

        # When rows move, sync ordering
        table.model().rowsMoved.connect(lambda *args, t=table: self.sync_model_order_from_totals_table(t))

    def sync_model_order_from_totals_table(self, table: QTableWidget):
        """
        Reads the Model column order from the (reordered) totals table and applies it globally.
        """
        order = []
        for r in range(table.rowCount()):
            it = table.item(r, 0)
            if it:
                m = it.text().strip()
                if m:
                    order.append(m)

        if not order:
            return

        self.model_order = list(dict.fromkeys(order))

        self.renumber_all_labels()
        self.rebuild_overlay_items()
        self.refresh_pages_list_ui()
        self.refresh_totals_ui()

        if self.active_highlight_model:
            self.apply_model_halos(self.active_highlight_model)


    def nearest_hue_distance_for_model(self, model: str) -> int:
        """
        Distance from this model's sample hue to the closest OTHER model hue.
        Used to auto-tighten thresholds for near-neighbor colors (like two yellows).
        """
        s = self.samples.get(model)
        if not s:
            return 180

        h0 = s.hsv[0]
        others = [ms.hsv[0] for m, ms in self.samples.items() if m != model]
        if not others:
            return 180

        return min(hue_distance(h0, h) for h in others)

    def model_specific_dh(self, model: str) -> int:
        """
        Auto-tighten hue window when neighbors are close.
        Example: if nearest neighbor is 8 hue units away, dh becomes ~3-4.
        """
        base = int(self.detect_dh)
        nearest = self.nearest_hue_distance_for_model(model)

        # If nothing close, keep your base tolerance.
        if nearest >= 20:
            return base

        # Tighten hard when colors are close.
        tightened = max(3, (nearest // 2) - 1)  # e.g. nearest=8 => dh=3
        return max(3, min(base, tightened))

    def predict_model_by_hue(self, hsv: Tuple[int, int, int]) -> Optional[str]:
        """
        Predict model by nearest hue distance among samples.
        """
        if not self.samples:
            return None
        best_model = None
        best_dist = 9999
        for model, sample in self.samples.items():
            d = hue_distance(hsv[0], sample.hsv[0])
            if d < best_dist:
                best_dist = d
                best_model = model
        return best_model

    # ---------------- Dot Data Ops ----------------

    def add_dot_at(self, page_index: int, x_px: float, y_px: float):
        """
        Smarter Add: sample ring color, predict model, allow override via dialog.
        """
        bgr = self.render_page_to_bgr(page_index)
        hsv_img = cv2.cvtColor(bgr, cv2.COLOR_BGR2HSV)

        ring = sample_dot_hsv(hsv_img, int(x_px), int(y_px))
        predicted = self.predict_model_by_hue(ring) if ring else None

        model_list = self.get_model_list()
        if not model_list:
            QMessageBox.information(self, "Add Dot", "No models exist yet. Pick at least one sample first.")
            return

        # Dialog: choose model
        dlg = QDialog(self)
        dlg.setWindowTitle("Add Dot")
        form = QFormLayout(dlg)

        cmb = QComboBox()
        cmb.addItems(model_list)
        if predicted and predicted in model_list:
            cmb.setCurrentText(predicted)
        form.addRow("Model", cmb)

        if ring:
            form.addRow("Sampled HSV", QLabel(str(ring)))
        else:
            form.addRow("Sampled HSV", QLabel("None (click closer to dot ring)"))

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(dlg.accept)
        buttons.rejected.connect(dlg.reject)
        form.addRow(buttons)

        if dlg.exec() != QDialog.Accepted:
            return

        model = cmb.currentText().strip()

        x_pdf, y_pdf = self.px_to_pdf(x_px, y_px)
        dot = Dot(
            page_index=page_index,
            cx_pdf=x_pdf,
            cy_pdf=y_pdf,
            radius_pdf=float(self.dot_r_expected_pdf),  # Feature 1: expected dot radius in PDF units
            model=model,
            excluded=False,
            sampled_hsv=ring
        )

        self.dots_by_page.setdefault(page_index, []).append(dot)

        if model not in self.model_order:
            self.model_order.append(model)
            self.refresh_model_order_ui()

        self.renumber_all_labels()
        self.rebuild_overlay_items()

        # NEW: Ensure totals + pages reflect the new dot immediately
        self.refresh_pages_list_ui()
        self.refresh_totals_ui()

        # If highlight is active, keep halos updated
        if self.active_highlight_model:
            self.apply_model_halos(self.active_highlight_model)

    def clear_current_page_dots(self):
        # figure out current page based on view center
        if not self.doc:
            return
        center = self.view.mapToScene(self.view.viewport().rect().center())
        page_index, _, _ = self.scene_to_page_px(center)
        if page_index is None:
            return
        self.dots_by_page[page_index] = []
        self.renumber_all_labels()
        self.rebuild_overlay_items()

    def get_model_list(self) -> List[str]:
        # prefer model_order, but include any sample models
        models = list(dict.fromkeys(self.model_order + list(self.samples.keys())))
        return [m for m in models if m.strip()]

    # ADD THIS: parse model tokens like C4, C11, M-1, etc.
    # REPLACE WITH THIS: parses "C4", "C 4", "C-4", "M-1", etc.
    def _parse_model_token(self, model: str) -> Tuple[str, Optional[int]]:
        """
        Returns (prefix, number) for models like:
          - "C4"   -> ("C", 4)
          - "C 4"  -> ("C", 4)
          - "C-4"  -> ("C", 4)
          - "M-1"  -> ("M", 1)
        If not parseable as <letters><optional-sep><digits>, returns (MODEL, None).
        """
        s = (model or "").strip().upper()
        m = re.match(r"^([A-Z]+)\s*[-_ ]?\s*(\d+)$", s)
        if not m:
            return (s, None)
        prefix = m.group(1)
        num = int(m.group(2))
        return (prefix, num)

    # ADD THIS: auto-insert newly-sampled model before higher-numbered siblings
    # REPLACE WITH THIS: prefix order first (A before C), then numeric within prefix (C4 before C11)
    def _auto_insert_model_order(self, model: str):
        """
        Inserts a NEW model into model_order using this precedence:
          1) Prefix letters (A < B < C ...)
          2) Numeric suffix within the prefix (4 < 11)
          3) Non-numeric tokens are treated as "after" numeric tokens of the same prefix.

        IMPORTANT: We do NOT reorder existing items; we only choose an insertion point for the new model.
        """
        model_u = (model or "").strip().upper()
        if not model_u:
            return
        if model_u in self.model_order:
            return

        new_prefix, new_num = self._parse_model_token(model_u)

        def sort_key(prefix: str, num: Optional[int]) -> Tuple[str, int, int]:
            # num_first: numeric tokens come before non-numeric tokens for same prefix
            num_first = 0 if num is not None else 1
            num_val = int(num) if num is not None else 10 ** 9
            return (prefix, num_val, num_first)

        new_key = sort_key(new_prefix, new_num)

        insert_at = None
        for i, existing in enumerate(self.model_order):
            ex_prefix, ex_num = self._parse_model_token(existing)
            ex_key = sort_key(ex_prefix, ex_num)

            # Insert BEFORE the first existing item whose key is "greater" than the new model.
            # This yields A* before C*, and C4 before C11.
            if new_key < ex_key:
                insert_at = i
                break

        if insert_at is None:
            self.model_order.append(model_u)
        else:
            self.model_order.insert(insert_at, model_u)

    def enter_text_select_mode(self, callback: Callable[[str], None]):
        """
        Arms a one-shot rectangle-drag mode on the map.
        When the user drags and releases, we extract PDF words inside the rect
        and call callback(text). Then we exit the mode.
        """
        self.exit_tool_modes()
        self.mode_text_select = True
        self.text_select_callback = callback
        self.text_select_drag_start_scene = None

        self.view.setDragMode(QGraphicsView.NoDrag)
        self.view.viewport().setCursor(Qt.CrossCursor)
        self.statusBar().showMessage("Text Select: click-drag a rectangle around nearby text, release to capture.")

    def extract_pdf_text_in_rect(self, page_index: int, rect_px: QRectF) -> str:
        """
        Use PyMuPDF’s word extraction (no OCR):
        - Convert selection rect from px -> PDF points
        - Return words inside, in reading-ish order
        """
        if not self.doc:
            return ""

        # rect_px is page-local px; convert to PDF coords
        x0_pdf, y0_pdf = self.px_to_pdf(rect_px.left(), rect_px.top())
        x1_pdf, y1_pdf = self.px_to_pdf(rect_px.right(), rect_px.bottom())
        sel = fitz.Rect(min(x0_pdf, x1_pdf), min(y0_pdf, y1_pdf), max(x0_pdf, x1_pdf), max(y0_pdf, y1_pdf))

        page = self.doc[page_index]
        words = page.get_text("words")  # (x0,y0,x1,y1,"word", block, line, word_no)

        picked = []
        for w in words:
            wx0, wy0, wx1, wy1, txt = w[0], w[1], w[2], w[3], w[4]
            wr = fitz.Rect(wx0, wy0, wx1, wy1)
            if wr.intersects(sel):
                picked.append((wy0, wx0, txt))

        picked.sort(key=lambda t: (t[0], t[1]))
        text = " ".join([t[2] for t in picked]).strip()
        return text

    # ---------------- Dot Edit UI ----------------

    def open_dot_edit_dialog(self, dot_id: int):
        page_i, dot_i = self.dot_index[dot_id]
        dot = self.dots_by_page[page_i][dot_i]

        dlg = DotEditDialog(self, dot, self.get_model_list())
        if dlg.exec() != QDialog.Accepted:
            return

        vals = dlg.result_values()
        if vals["delete"]:
            # remove dot
            self.dots_by_page[page_i].pop(dot_i)
        else:
            dot.model = vals["model"]
            dot.excluded = vals["excluded"]

        if dot.model and dot.model not in self.model_order:
            self.model_order.append(dot.model)
            self.refresh_model_order_ui()

        self.renumber_all_labels()
        self.rebuild_overlay_items()
    def _iter_all_visible_dots(self) -> List[Dot]:
        dots: List[Dot] = []
        for pi in self.visible_pages:
            dots.extend(self.dots_by_page.get(pi, []))
        return [d for d in dots if not d.excluded and d.label]

    def _label_number(self, label: str) -> int:
        # label format: "{n}-{model}"
        try:
            return int(label.split("-", 1)[0])
        except Exception:
            return 999999

    def get_all_dots_sorted_by_number(self) -> List[Dot]:
        dots = self._iter_all_visible_dots()
        dots.sort(key=lambda d: self._label_number(d.label or "999999-"))
        return dots
    def highlight_names_uid(self, uid: Optional[str]):
        """
        Ensures NAMES tab is selected, highlights the row for uid,
        and scrolls so that row appears centered.
        """
        self.names_highlight_uid = uid

        if not uid:
            return

        # Find the NAMES tab index by title
        names_tab_index = -1
        for i in range(self.totals_tabs.count()):
            if self.totals_tabs.tabText(i) == "NAMES":
                names_tab_index = i
                break

        if names_tab_index < 0:
            return

        # Force NAMES tab selection if requested (or if naming is active)
        if self.force_names_tab_selected:
            self.totals_tabs.setCurrentIndex(names_tab_index)

        table = self.totals_tabs.widget(names_tab_index)
        if not isinstance(table, QTableWidget):
            return

        # Find the row that stores this uid
        for r in range(table.rowCount()):
            it0 = table.item(r, 0)
            if not it0:
                continue
            row_uid = it0.data(Qt.UserRole)
            if row_uid == uid:
                table.setCurrentCell(r, 0)
                table.selectRow(r)
                table.scrollToItem(it0, QAbstractItemView.PositionAtCenter)
                break

    def make_names_table_widget(self) -> QTableWidget:
        """
        Builds the NAMES tab table: Number | Camera Type | Name
        Reflects current dot labels + names.
        """
        dots = self.get_all_dots_sorted_by_number()

        table = QTableWidget(0, 3)
        table.setHorizontalHeaderLabels(["Number", "Camera Type", "Name"])
        table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        table.verticalHeader().setVisible(False)
        table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        table.setSelectionBehavior(QAbstractItemView.SelectRows)
        table.setSelectionMode(QAbstractItemView.SingleSelection)

        for d in dots:
            row = table.rowCount()
            table.insertRow(row)

            num = self._label_number(d.label or "")

            it_num = QTableWidgetItem(str(num))
            it_num.setData(Qt.UserRole, d.uid)  # <-- IMPORTANT: row -> dot uid

            table.setItem(row, 0, it_num)
            table.setItem(row, 1, QTableWidgetItem(d.model))
            table.setItem(row, 2, QTableWidgetItem(d.name or ""))

        # Clicking a row in NAMES: focus that single dot + single breathing halo
        table.cellClicked.connect(lambda r, c, t=table: self.on_names_row_clicked(t, r))

        return table

    # ---------------- Detection Pipeline ----------------

    def render_page_to_bgr(self, page_index: int) -> np.ndarray:
        """
        Cached render. Rendering a page pixmap is expensive; the magnifier calls this a lot.
        Cache key includes render_zoom because zoom changes the raster.
        """
        if not self.doc:
            return np.zeros((1, 1, 3), dtype=np.uint8)

        key = (int(page_index), float(self.render_zoom))
        cached = self._page_bgr_cache.get(key)
        if cached is not None:
            return cached

        page = self.doc[page_index]
        mat = fitz.Matrix(self.render_zoom, self.render_zoom)
        pix = page.get_pixmap(matrix=mat, alpha=False)
        rgb = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.height, pix.width, 3)
        bgr = cv2.cvtColor(rgb, cv2.COLOR_RGB2BGR)

        self._page_bgr_cache[key] = bgr
        return bgr

    def render_page_to_hsv(self, page_index: int) -> np.ndarray:
        """
        Cached HSV render. Avoid converting per-frame in update_magnifier().
        """
        if not self.doc:
            return np.zeros((1, 1, 3), dtype=np.uint8)

        key = (int(page_index), float(self.render_zoom))
        cached = self._page_hsv_cache.get(key)
        if cached is not None:
            return cached

        bgr = self.render_page_to_bgr(page_index)
        hsv = cv2.cvtColor(bgr, cv2.COLOR_BGR2HSV)

        self._page_hsv_cache[key] = hsv
        return hsv

    def match_gate(self, model: str, ring_hsv: Tuple[int, int, int]) -> bool:
        """
        Extra safety gate so "close-ish blue" doesn't become the wrong model.
        Stronger sample-relative S/V filtering to prevent C1 vs C14 confusion.
        """
        sample = self.samples.get(model)
        if not sample:
            return True

        h0, s0, v0 = sample.hsv
        h, s, v = ring_hsv

        # Hue must be close
        if hue_distance(h, h0) > self.detect_max_hue_dist:
            return False

        # Strong anti-washout rules:
        # if the sample was vivid, don't accept dull versions
        if s0 >= 80:
            if s < int(s0 * 0.70):
                return False

        # avoid very faint / near-white hits
        if v < int(v0 * 0.55):
            return False

        # still allow some variation
        if abs(s - s0) > self.detect_max_sat_diff:
            return False
        if abs(v - v0) > self.detect_max_val_diff:
            return False

        return True

    def model_match_score(self, model: str, ring_hsv: Tuple[int, int, int]) -> float:
        """
        Lower score = better match.
        Weighted HSV distance from the stored sample HSV.
        """
        sample = self.samples.get(model)
        if not sample:
            return 999999.0

        h0, s0, v0 = sample.hsv
        h, s, v = ring_hsv

        hd = hue_distance(h, h0)
        sd = abs(s - s0)
        vd = abs(v - v0)

        # Hue matters most; S/V help separate vivid vs washed-out lookalikes
        return (hd * 3.0) + (sd * 0.60) + (vd * 0.45)

    def best_model_for_ring(self, ring_hsv: Tuple[int, int, int]) -> Tuple[Optional[str], float, float]:
        """
        Returns: (best_model, best_score, second_best_score)
        Used to prevent cross-model stealing (ex: C4 picking up C11 blobs).
        """
        if not self.samples:
            return None, 999999.0, 999999.0

        scored = []
        for m in self.samples.keys():
            scored.append((self.model_match_score(m, ring_hsv), m))

        scored.sort(key=lambda x: x[0])
        best_score, best_model = scored[0]
        second_score = scored[1][0] if len(scored) > 1 else 999999.0
        return best_model, best_score, second_score

    def dedupe_dots_px(self, dots_px, dist_thresh_px: float):
        """
        Dedupe close detections: keep the BEST model match score, not "largest radius".

        dots_px: [(cx, cy, r, model, ringHSV)]
        """
        kept = []  # (cx, cy, r, model, ringHSV, score)

        for cx, cy, r, model, ring in dots_px:
            score = self.model_match_score(model, ring)

            placed = False
            for i, (kx, ky, kr, km, kring, kscore) in enumerate(kept):
                dist = math.hypot(cx - kx, cy - ky)
                if dist < dist_thresh_px:
                    # Same physical dot — keep the better model
                    if score < kscore:
                        kept[i] = (cx, cy, r, model, ring, score)
                    placed = True
                    break

            if not placed:
                kept.append((cx, cy, r, model, ring, score))

        # Return without score field
        return [(cx, cy, r, model, ring) for (cx, cy, r, model, ring, score) in kept]

    def generate_preview_current_page(self):
        if not self.doc:
            return
        if not self.samples or not self.model_order:
            QMessageBox.information(self, "Preview", "Pick at least one model sample first.")
            return

        center = self.view.mapToScene(self.view.viewport().rect().center())
        page_index, _, _ = self.scene_to_page_px(center)
        if page_index is None:
            page_index = 0

        self.detect_page(page_index)
        self.renumber_all_labels()
        self.rebuild_overlay_items()
        self.refresh_pages_list_ui()
        self.refresh_totals_ui()


    def generate_preview_all_pages(self):
        if not self.doc:
            return
        if not self.samples or not self.model_order:
            QMessageBox.information(self, "Preview", "Pick at least one model sample first.")
            return

        for page_index in self.visible_pages:
            self.detect_page(page_index)

        self.renumber_all_labels()
        self.rebuild_overlay_items()
        self.refresh_pages_list_ui()
        self.refresh_totals_ui()


    def detect_page(self, page_index: int):
        """
        Detect dots for each model with a color sample. Store dots in PDF coords.
        """
        bgr = self.render_page_to_bgr(page_index)

        detections_px = []  # (cx, cy, r, model, ringHSV)
        for model in self.model_order:
            if model not in self.samples:
                continue
            # Feature 1: derive area range from dot radius min/max (PDF pts -> px via render_zoom)
            r_min_px = max(1.0, float(self.dot_r_min_pdf) * float(self.render_zoom))
            r_max_px = max(r_min_px, float(self.dot_r_max_pdf) * float(self.render_zoom))

            area_min = int(math.pi * (r_min_px ** 2))
            area_max = int(math.pi * (r_max_px ** 2))

            hits = detect_dots_for_model(
                bgr=bgr,
                model_sample=self.samples[model],
                area_min=area_min,
                area_max=area_max,
                circ_min=self.detect_circ_min,
                dh=self.model_specific_dh(model),
                min_s=self.detect_min_s,
                min_v=self.detect_min_v
            )

            for cx, cy, r, ring in hits:
                # reject things that match hue-mask but are too far from the true sample
                if not self.match_gate(model, ring):
                    continue

                # FINAL SAFETY: Only keep this detection if THIS model is actually the best match
                best_model, best_score, second_score = self.best_model_for_ring(ring)

                # If some other model matches better, don't allow this model to "steal" the blob
                if best_model is None or best_model != model:
                    continue

                # Optional confidence margin (helps close-hue cases like yellow vs lime)
                # If best and second-best are too close, skip the detection as ambiguous
                # If this model has a close hue neighbor (ex: two yellows), require a bigger margin.
                nearest = self.nearest_hue_distance_for_model(model)
                required_margin = 18 if nearest >= 15 else 32

                if (second_score - best_score) < required_margin:
                    continue

                detections_px.append((cx, cy, r, model, ring))

        detections_px = self.dedupe_dots_px(detections_px, dist_thresh_px=self.dedupe_dist_px)
        # Filter out detections inside exclusion zones (stored in PDF coords)
        zones_pdf = self.exclusion_zones_by_page.get(page_index, [])
        if zones_pdf:
            zones_px = []
            for z in zones_pdf:
                x1 = z.left() * self.render_zoom
                y1 = z.top() * self.render_zoom
                x2 = z.right() * self.render_zoom
                y2 = z.bottom() * self.render_zoom
                zones_px.append((min(x1, x2), min(y1, y2), max(x1, x2), max(y1, y2)))

            filtered = []
            for cx, cy, r, model, ring in detections_px:
                inside = False
                for (zx1, zy1, zx2, zy2) in zones_px:
                    if zx1 <= cx <= zx2 and zy1 <= cy <= zy2:
                        inside = True
                        break
                if not inside:
                    filtered.append((cx, cy, r, model, ring))
            detections_px = filtered

        dots_pdf: List[Dot] = []
        for cx, cy, r, model, ring in detections_px:
            x_pdf, y_pdf = self.px_to_pdf(cx, cy)
            dot = Dot(
                page_index=page_index,
                cx_pdf=x_pdf,
                cy_pdf=y_pdf,
                radius_pdf=(r / self.render_zoom),
                model=model,
                excluded=False,
                label_dx_pdf=6.0,
                label_dy_pdf=-6.0,
                sampled_hsv=ring
            )
            dots_pdf.append(dot)

        # Preserve uid/name from existing dots on this page (best-effort match)
        prev = self.dots_by_page.get(page_index, [])
        used_prev = set()

        def best_prev_match(new_dot: Dot) -> Optional[Dot]:
            best = None
            best_dist = 1e9
            for pd in prev:
                if pd.excluded:
                    continue
                if pd.model != new_dot.model:
                    continue
                if pd.uid in used_prev:
                    continue
                dx = pd.cx_pdf - new_dot.cx_pdf
                dy = pd.cy_pdf - new_dot.cy_pdf
                dist = math.hypot(dx, dy)
                # threshold in PDF points (~10 px at your default zoom 2.5)
                if dist < best_dist and dist <= (10.0 / self.render_zoom):
                    best = pd
                    best_dist = dist
            return best

        for nd in dots_pdf:
            match = best_prev_match(nd)
            if match:
                nd.uid = match.uid
                nd.name = match.name
                used_prev.add(match.uid)

        self.dots_by_page[page_index] = dots_pdf

    def remove_dots_inside_exclusion_rect(self, page_index: int, rect_pdf: QRectF) -> int:
        """
        Remove any existing dots whose centers fall inside rect_pdf on page_index.
        Returns number removed.
        """
        dots = self.dots_by_page.get(page_index, [])
        if not dots:
            return 0

        kept = []
        removed = 0

        for d in dots:
            if rect_pdf.contains(QPointF(d.cx_pdf, d.cy_pdf)):
                removed += 1
            else:
                kept.append(d)

        self.dots_by_page[page_index] = kept
        return removed

    def delete_dot_by_id(self, dot_id: int) -> bool:
        """
        Delete one dot (and its label/ellipse overlay) by dot_id.
        Returns True if a dot was removed.
        """
        try:
            if dot_id is None:
                return False
            if dot_id < 0 or dot_id >= len(self.dot_index):
                return False

            page_index, dot_i = self.dot_index[dot_id]
            dots = self.dots_by_page.get(page_index, [])
            if dot_i < 0 or dot_i >= len(dots):
                return False

            # Remove from the data model
            dots.pop(dot_i)
            if not dots:
                # Optional: keep dict clean
                self.dots_by_page.pop(page_index, None)

            # Renumber first (works off dots_by_page), then rebuild overlays
            self.renumber_all_labels()
            self.rebuild_overlay_items()
            self.refresh_pages_list_ui()
            self.refresh_totals_ui()

            return True
        except Exception:
            return False

    # ---------------- Numbering / Labels ----------------

    def renumber_all_labels(self):
        """
        Assign labels like 1-C1, 2-C1, 3-C1, 4-C3, 5-C3, 6-C4 ... etc.

        Key behavior:
        - Numbering continues across models in self.model_order (sampling order).
        - Reading order inside each model group stays top-to-bottom, then left-to-right.
        - Numbering continues across ALL pages (Page 2 starts after Page 1 ends).
        """
        if not self.doc:
            return

        running = 1  # <-- global counter ACROSS ALL PAGES

        for page_index in self.visible_pages:
            dots = self.dots_by_page.get(page_index, [])

            # wipe old label strings
            for d in dots:
                d.label = None

            for model in self.model_order:
                group = [d for d in dots if (not d.excluded and d.model == model)]
                group.sort(key=lambda x: (x.cy_pdf, x.cx_pdf))

                for d in group:
                    d.label = f"{running}-{model}"
                    running += 1

        # If labels moved (draggable), keep their offsets intact

    # ---------------- Overlay Rebuild ----------------

    def rebuild_overlay_items(self):
        """
        Rebuild dot ellipse overlays + label overlays for ALL pages.
        This keeps the scene synchronized after edits/detection.
        """
        # Remove existing overlay items
        # Remove existing overlay items
        for it in list(self.dot_ellipse_items.values()):
            self.scene.removeItem(it)
        for it in list(self.dot_label_items.values()):
            self.scene.removeItem(it)
        for it in list(self.dot_connector_items.values()):
            self.scene.removeItem(it)

        self.dot_ellipse_items.clear()
        self.dot_label_items.clear()
        self.dot_connector_items.clear()
        self.dot_index.clear()

        # Recreate dot overlays
        dot_id = 0

        for page_index, dots in sorted(self.dots_by_page.items(), key=lambda x: x[0]):
            for dot_i, d in enumerate(dots):
                self.dot_index.append((page_index, dot_i))

                if d.excluded:
                    dot_id += 1
                    continue

                cx_px, cy_px = self.pdf_to_px(d.cx_pdf, d.cy_pdf)
                rx_px = max(6.0, d.radius_pdf * self.render_zoom)
                scene_pos = self.page_px_to_scene(page_index, cx_px, cy_px)

                # dot ellipse (ring color = slightly darker sampled model color)
                ring_c = self._darker_qcolor(self.model_color(d.model), 20)
                pen = QPen(ring_c, 2)

                ellipse = DotEllipseItem(
                    dot_id,
                    scene_pos.x() - rx_px,
                    scene_pos.y() - rx_px,
                    rx_px * 2,
                    rx_px * 2
                )
                ellipse.setPen(pen)
                ellipse.setBrush(Qt.NoBrush)
                ellipse.setZValue(10)
                self.scene.addItem(ellipse)
                self.dot_ellipse_items[dot_id] = ellipse

                # label
                if d.label:
                    model_c = self.model_color(d.model)

                    label_item = LabelTextItem(
                        dot_id,
                        d.label,
                        base_font_pt=self.base_label_font_pt,
                        scale=self.label_scale_factor(),
                        border_color=model_c,
                        on_moved=self.update_connector_for_dot
                    )

                    # compute label scene position from dot + offset (PDF->px->scene)
                    lx_pdf = d.cx_pdf + d.label_dx_pdf
                    ly_pdf = d.cy_pdf + d.label_dy_pdf
                    lx_px, ly_px = self.pdf_to_px(lx_pdf, ly_pdf)
                    label_scene = self.page_px_to_scene(page_index, lx_px, ly_px)
                    label_item.setPos(label_scene)

                    # White background (simple approach: draw a small rect behind with selection)
                    label_item.setZValue(20)
                    self.scene.addItem(label_item)
                    self.dot_label_items[dot_id] = label_item

                    self.update_connector_for_dot(dot_id)

                dot_id += 1

        # Auto-pack labels so they don't overlap other labels or dots
        self.resolve_label_collisions()

        # Persist the new positions back into dot.label_dx_pdf / label_dy_pdf
        self.update_label_offsets_from_scene()

        self.statusBar().showMessage("Overlay rebuilt. Labels are draggable.")

    def find_dot_item_near_scene(self, pos_scene: QPointF, max_dist: float = 18) -> Optional[int]:
        """
        Find dot_id nearest to pos_scene among current ellipse items.
        """
        best = None
        best_d = 1e9
        x, y = float(pos_scene.x()), float(pos_scene.y())

        for dot_id, ellipse in self.dot_ellipse_items.items():
            rect = ellipse.rect()
            cx = rect.x() + rect.width() / 2.0
            cy = rect.y() + rect.height() / 2.0
            d = math.hypot(cx - x, cy - y)
            if d < best_d and d <= max_dist:
                best_d = d
                best = dot_id

        return best
    def move_dot_to_scene_pos(self, dot_id: int, new_scene_pos: QPointF):
        """
        Move a dot by dragging in the scene.
        Updates dot PDF coords, then updates ellipse + label item positions.
        """
        if dot_id not in self.dot_ellipse_items:
            return
        if dot_id >= len(self.dot_index):
            return

        page_i, dot_i = self.dot_index[dot_id]
        dot = self.dots_by_page[page_i][dot_i]

        # Convert scene position -> page pixel coords
        page_index, x_px, y_px = self.scene_to_page_px(new_scene_pos)
        if page_index is None:
            return

        # Prevent dragging dot onto a different page accidentally:
        if page_index != dot.page_index:
            return

        # Update dot PDF coords
        x_pdf, y_pdf = self.px_to_pdf(x_px, y_px)
        dot.cx_pdf = x_pdf
        dot.cy_pdf = y_pdf

        # Update ellipse item geometry in the scene
        cx_px, cy_px = self.pdf_to_px(dot.cx_pdf, dot.cy_pdf)
        rx_px = max(6.0, dot.radius_pdf * self.render_zoom)
        dot_scene = self.page_px_to_scene(dot.page_index, cx_px, cy_px)

        ellipse = self.dot_ellipse_items[dot_id]
        ellipse.setRect(dot_scene.x() - rx_px, dot_scene.y() - rx_px, rx_px * 2, rx_px * 2)

        # Update label position to follow dot + stored offsets
        label_item = self.dot_label_items.get(dot_id)
        if label_item:
            lx_pdf = dot.cx_pdf + dot.label_dx_pdf
            ly_pdf = dot.cy_pdf + dot.label_dy_pdf
            lx_px, ly_px = self.pdf_to_px(lx_pdf, ly_pdf)
            label_scene = self.page_px_to_scene(dot.page_index, lx_px, ly_px)
            label_item.setPos(label_scene)
        self.update_connector_for_dot(dot_id)


    # Update offsets after label drag: call from a timer, a button, or on demand
    def update_label_offsets_from_scene(self):
        """
        If user drags labels, update stored dot.label_dx_pdf/dy_pdf based on new label scene pos.
        """
        for dot_id, label_item in self.dot_label_items.items():
            page_i, dot_i = self.dot_index[dot_id]
            dot = self.dots_by_page[page_i][dot_i]

            # label scene position -> page px -> pdf
            label_scene_pos = label_item.pos()
            page_index, x_px, y_px = self.scene_to_page_px(label_scene_pos)
            if page_index is None:
                continue

            # dot px
            dot_x_px, dot_y_px = self.pdf_to_px(dot.cx_pdf, dot.cy_pdf)

            # offsets in px
            dx_px = x_px - dot_x_px
            dy_px = y_px - dot_y_px

            # store in pdf units
            dot.label_dx_pdf = dx_px / self.render_zoom
            dot.label_dy_pdf = dy_px / self.render_zoom
        self.rebuild_all_connectors()

    # ---------------- Spatial Grid Helpers (for fast local collision checks) ----------------

    def _grid_cell(self, x: float, y: float, cell_size: float) -> Tuple[int, int]:
        """Return integer grid cell coordinates for a point."""
        return (int(math.floor(x / cell_size)), int(math.floor(y / cell_size)))

    def _grid_keys_for_rect(self, rect: QRectF, cell_size: float) -> List[Tuple[int, int]]:
        """Return all grid cell keys overlapped by rect."""
        x0 = rect.left()
        y0 = rect.top()
        x1 = rect.right()
        y1 = rect.bottom()

        gx0, gy0 = self._grid_cell(x0, y0, cell_size)
        gx1, gy1 = self._grid_cell(x1, y1, cell_size)

        keys = []
        for gy in range(gy0, gy1 + 1):
            for gx in range(gx0, gx1 + 1):
                keys.append((gx, gy))
        return keys

    def _grid_neighbor_keys(self, gx: int, gy: int, radius_cells: int) -> List[Tuple[int, int]]:
        """Return neighboring grid keys within radius_cells (Chebyshev neighborhood)."""
        out = []
        for yy in range(gy - radius_cells, gy + radius_cells + 1):
            for xx in range(gx - radius_cells, gx + radius_cells + 1):
                out.append((xx, yy))
        return out


    def _circle_intersects_rect(self, cx: float, cy: float, r: float, rect: QRectF) -> bool:
        """
        True if circle (cx,cy,r) intersects rect (scene coords).
        """
        # nearest point on rect to circle center
        nx = min(max(cx, rect.left()), rect.right())
        ny = min(max(cy, rect.top()), rect.bottom())
        dx = cx - nx
        dy = cy - ny
        return (dx * dx + dy * dy) <= (r * r)

    def resolve_label_collisions(
            self,
            max_iter: int = 120,
            step_px: float = 6.0,
            dot_padding_px: float = 6.0,
            label_padding_px: float = 6.0,
            grid_cell_px: float = 220.0,
            neighbor_cells: int = 1
    ):
        """
        Push label boxes so they don't overlap each other or any dot circle.

        OPTIMIZED:
        - Uses a uniform grid spatial index so each label only checks nearby labels/dots.
        - grid_cell_px controls locality size (bigger => more neighbors checked, safer but slower).
        - neighbor_cells controls how many cells around the label we consider (1 => 3x3).
        """
        if not self.dot_label_items:
            return

        # ---------------------------
        # Build DOT obstacles + DOT grid (static during packing)
        # ---------------------------
        dot_obs = []  # list of (cx, cy, r) in SCENE coords
        dot_grid: Dict[Tuple[int, int], List[int]] = {}  # cell -> list of dot indices

        for dot_id, ellipse in self.dot_ellipse_items.items():
            er = ellipse.rect()
            dcx = er.x() + er.width() / 2.0
            dcy = er.y() + er.height() / 2.0
            dr = (er.width() / 2.0) + dot_padding_px
            dot_index = len(dot_obs)
            dot_obs.append((dcx, dcy, dr))

            # Put dot into grid cells it overlaps (use bounding rect of circle)
            dot_rect = QRectF(dcx - dr, dcy - dr, dr * 2.0, dr * 2.0)
            for key in self._grid_keys_for_rect(dot_rect, grid_cell_px):
                dot_grid.setdefault(key, []).append(dot_index)

        label_ids = list(self.dot_label_items.keys())

        # Helper: label rect in scene coords with padding
        def label_rect(dot_id: int) -> QRectF:
            li = self.dot_label_items[dot_id]
            r = li.mapToScene(li.boundingRect()).boundingRect()
            return r.adjusted(-label_padding_px, -label_padding_px, label_padding_px, label_padding_px)

        # ---------------------------
        # Iterate & push (rebuild label grid each iter)
        # ---------------------------
        for _ in range(max_iter):
            moved_any = False

            # Build rects + label grid for this iteration
            rects: Dict[int, QRectF] = {}
            label_grid: Dict[Tuple[int, int], List[int]] = {}

            for i in label_ids:
                ri = label_rect(i)
                rects[i] = ri
                for key in self._grid_keys_for_rect(ri, grid_cell_px):
                    label_grid.setdefault(key, []).append(i)

            # Push each label away from nearby dots/labels
            for i in label_ids:
                li = self.dot_label_items[i]
                ri = rects[i]
                ci = ri.center()

                push_x = 0.0
                push_y = 0.0

                # Determine which grid cell the label center is in
                gx, gy = self._grid_cell(ci.x(), ci.y(), grid_cell_px)

                # Collect candidate dot indices from neighboring cells
                cand_dot_indices: List[int] = []
                for nk in self._grid_neighbor_keys(gx, gy, neighbor_cells):
                    cand_dot_indices.extend(dot_grid.get(nk, []))

                # Collect candidate label ids from neighboring cells
                cand_label_ids: List[int] = []
                for nk in self._grid_neighbor_keys(gx, gy, neighbor_cells):
                    cand_label_ids.extend(label_grid.get(nk, []))

                # (Optional) dedupe candidates cheaply
                if cand_dot_indices:
                    cand_dot_indices = list(set(cand_dot_indices))
                if cand_label_ids:
                    cand_label_ids = list(set(cand_label_ids))

                # 1) avoid DOTS (local)
                for di in cand_dot_indices:
                    dcx, dcy, dr = dot_obs[di]
                    if self._circle_intersects_rect(dcx, dcy, dr, ri):
                        vx = ci.x() - dcx
                        vy = ci.y() - dcy
                        mag = math.hypot(vx, vy) or 1.0
                        push_x += (vx / mag)
                        push_y += (vy / mag)

                # 2) avoid OTHER LABELS (local)
                for j in cand_label_ids:
                    if j == i:
                        continue
                    rj = rects.get(j)
                    if rj is None:
                        continue
                    if ri.intersects(rj):
                        cj = rj.center()
                        vx = ci.x() - cj.x()
                        vy = ci.y() - cj.y()
                        mag = math.hypot(vx, vy) or 1.0
                        push_x += (vx / mag)
                        push_y += (vy / mag)

                if push_x != 0.0 or push_y != 0.0:
                    mag = math.hypot(push_x, push_y) or 1.0
                    dx = (push_x / mag) * step_px
                    dy = (push_y / mag) * step_px

                    li.setPos(li.pos() + QPointF(dx, dy))
                    moved_any = True

            if not moved_any:
                break

        # After moving labels, rebuild connectors so they attach to the new locations
        self.rebuild_all_connectors()

    # ---------------- Exports ----------------

    def export_labeled_pdf(self):
        if not self.doc:
            return

        # ensure we store label drags
        self.update_label_offsets_from_scene()

        out_path, _ = QFileDialog.getSaveFileName(self, "Save Labeled PDF", "", "PDF Files (*.pdf)")
        if not out_path:
            return

        try:
            out = fitz.open()
            for page_index in self.visible_pages:
                # Clone the original page into the output PDF (preserves annotations/layers better)
                out.insert_pdf(self.doc, from_page=page_index, to_page=page_index, annots=True, links=True)
                new_page = out[-1]  # the page we just inserted

                dots = self.dots_by_page.get(page_index, [])
                for d in dots:
                    if d.excluded or not d.label:
                        continue

                    # label position in PDF coords
                    lx = d.cx_pdf + d.label_dx_pdf
                    ly = d.cy_pdf + d.label_dy_pdf

                    # background box
                    # estimate box width from label length (rough but effective)
                    scale = self.label_scale_factor()
                    font_pt = max(1.0, self.base_export_pdf_font_pt * scale)

                    # Sample color -> RGB 0..1 for PyMuPDF
                    c = self.model_color(d.model)
                    stroke = (c.red() / 255.0, c.green() / 255.0, c.blue() / 255.0)
                    stroke_w = max(0.5, 1.2 * scale)

                    # Estimate label box
                    box_w = max(40.0 * scale, (font_pt * 0.62) * len(d.label) + (10.0 * scale))
                    rect = fitz.Rect(
                        lx - (2.0 * scale),
                        ly - (font_pt + (2.0 * scale)),
                        lx + box_w,
                        ly + (4.0 * scale)
                    )

                    # Connector: from dot perimeter to nearest point on rect
                    dot_center = fitz.Point(d.cx_pdf, d.cy_pdf)

                    # nearest point on rect to dot center (clamp)
                    ax = min(max(dot_center.x, rect.x0), rect.x1)
                    ay = min(max(dot_center.y, rect.y0), rect.y1)
                    label_attach = fitz.Point(ax, ay)

                    # perimeter point on dot towards label_attach
                    dx = label_attach.x - dot_center.x
                    dy = label_attach.y - dot_center.y
                    dist = math.hypot(dx, dy) or 1.0
                    ux, uy = dx / dist, dy / dist
                    dot_r = max(1.0, float(d.radius_pdf))
                    dot_attach = fitz.Point(dot_center.x + ux * dot_r, dot_center.y + uy * dot_r)

                    new_page.draw_line(dot_attach, label_attach, color=stroke, width=stroke_w, overlay=True)

                    # White fill BELOW original content so colored dots remain visible
                    # Draw white fill UNDER original content so we don't erase underlying markers
                    new_page.draw_rect(rect, color=None, fill=(1, 1, 1), fill_opacity=0.5, width=0, overlay=False)

                    # Draw the colored border ON TOP so it always shows
                    new_page.draw_rect(rect, color=stroke, fill=None, width=stroke_w, overlay=True)

                    # Text over everything
                    new_page.insert_text((lx, ly), d.label, fontsize=font_pt, color=(0, 0, 0), overlay=True)

            out.save(out_path)
            out.close()
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to export labeled PDF:\n{e}")
            return

        QMessageBox.information(self, "Export", f"Saved labeled PDF:\n{out_path}")

    def on_names_row_clicked(self, table: QTableWidget, row: int):
        """
        Clicking a row in the NAMES tab should:
        - clear model-wide halos
        - focus and zoom to the single dot
        - show ONLY that dot's breathing halo
        """
        it = table.item(row, 0)
        if not it:
            return

        uid = it.data(Qt.UserRole)
        if not uid:
            return

        # Turn off model-wide highlight halos (ALL tab behavior)
        self.active_highlight_model = None
        self.clear_halos()

        # Focus one dot + zoom + breathing halo
        self.focus_dot(uid, pad_px=350.0)

        # Optional: keep the row visibly selected
        table.selectRow(row)

    def focus_dot_zoomed(self, uid: str, pad_px: float = 350.0):
        """
        SHOW CONTEXT:
        - Pan/zoom so the dot is centered and a readable surrounding area is visible.
        - Apply ONLY the single-dot breathing halo (no model-wide halos).
        """
        d = self.get_dot_by_uid(uid)
        if not d or d.excluded:
            return

        # Remember resume cursor for naming workflow
        self.naming_last_uid = uid

        # Convert dot center to scene coords
        cx_px, cy_px = self.pdf_to_px(d.cx_pdf, d.cy_pdf)
        scene_pos = self.page_px_to_scene(d.page_index, cx_px, cy_px)

        # Build a context rect around the dot in SCENE coords
        rect = QRectF(
            scene_pos.x() - pad_px,
            scene_pos.y() - pad_px,
            pad_px * 2,
            pad_px * 2
        )

        # Zoom to context + center on dot
        self.view.fitInView(rect, Qt.KeepAspectRatio)
        self.view.centerOn(scene_pos)

        # Apply single-dot breathing halo
        self.apply_focus_halo_to_dot(d)

    def export_images(self):
        if not self.doc:
            return

        # ensure we store label drags
        self.update_label_offsets_from_scene()

        folder = QFileDialog.getExistingDirectory(self, "Select Output Folder")
        if not folder:
            return

        dlg = ExportImagesDialog(self)
        if dlg.exec() != QDialog.Accepted:
            return
        dpi = dlg.dpi()

        # render scale
        zoom = dpi / 72.0

        try:
            for page_index in self.visible_pages:
                page = self.doc[page_index]
                mat = fitz.Matrix(zoom, zoom)
                pix = page.get_pixmap(matrix=mat, alpha=False)

                rgb = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.height, pix.width, 3)
                bgr = cv2.cvtColor(rgb, cv2.COLOR_RGB2BGR)

                # draw labels into the high-res image
                dots = self.dots_by_page.get(page_index, [])
                for d in dots:
                    if d.excluded or not d.label:
                        continue

                    # Convert PDF coords -> image px at this zoom
                    cx_px = int(d.cx_pdf * zoom)
                    cy_px = int(d.cy_pdf * zoom)
                    lx_px = int((d.cx_pdf + d.label_dx_pdf) * zoom)
                    ly_px = int((d.cy_pdf + d.label_dy_pdf) * zoom)

                    # draw white box + text
                    text = d.label
                    scale = self.label_scale_factor()

                    font = cv2.FONT_HERSHEY_SIMPLEX
                    font_scale = self.base_export_img_font_scale * scale
                    thickness = max(1, int(round(2 * scale)))
                    (tw, th), baseline = cv2.getTextSize(text, font, font_scale, thickness)

                    x1, y1 = lx_px - 2, ly_px - th - 2
                    x2, y2 = lx_px + tw + 4, ly_px + 4
                    # Color (sample) for border + connector, in BGR for OpenCV
                    qc = self.model_color(d.model)
                    bgr_col = (int(qc.blue()), int(qc.green()), int(qc.red()))

                    # Connector endpoints:
                    # - label attach = nearest point on label rect to dot center
                    dcx = int(d.cx_pdf * zoom)
                    dcy = int(d.cy_pdf * zoom)
                    dr = max(1, int(d.radius_pdf * zoom))

                    # Clamp dot center to label rect for nearest point
                    ax = min(max(dcx, x1), x2)
                    ay = min(max(dcy, y1), y2)

                    # Dot perimeter point towards (ax, ay)
                    dx = ax - dcx
                    dy = ay - dcy
                    dist = int(math.hypot(dx, dy)) or 1
                    ux = dx / float(dist)
                    uy = dy / float(dist)
                    px = int(dcx + ux * dr)
                    py = int(dcy + uy * dr)

                    # Draw connector line
                    cv2.line(bgr, (px, py), (ax, ay), bgr_col, thickness=max(1, int(round(2 * scale))), lineType=cv2.LINE_AA)

                    x1 = max(0, x1); y1 = max(0, y1)
                    x2 = min(bgr.shape[1]-1, x2); y2 = min(bgr.shape[0]-1, y2)

                    # Semi-transparent white background so original colored dots stay visible beneath
                    overlay = bgr.copy()
                    cv2.rectangle(bgr, (x1, y1), (x2, y2), (255, 255, 255), -1)
                    cv2.rectangle(bgr, (x1, y1), (x2, y2), bgr_col, thickness=max(1, int(round(2 * scale))),
                                  lineType=cv2.LINE_AA)
                    alpha = 0.75
                    bgr = cv2.addWeighted(overlay, alpha, bgr, 1 - alpha, 0)

                    cv2.putText(bgr, text, (lx_px, ly_px), font, font_scale, (0, 0, 0), thickness, cv2.LINE_AA)

                out_file = f"{folder}/page_{page_index+1:03d}_{dpi}dpi.png"
                cv2.imwrite(out_file, bgr)

        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to export images:\n{e}")
            return

        QMessageBox.information(self, "Export", f"Images exported to:\n{folder}")


# ---------------------------
# Run
# ---------------------------

def main():
    app = QApplication(sys.argv)
    w = PdfDotLabeler()
    w.show()
    sys.exit(app.exec())

if __name__ == "__main__":
    main()
